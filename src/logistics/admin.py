import io
import json
import time
from urllib.parse import urlencode
import zipfile

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.contrib.admin import helpers
from django.contrib.admin.views.main import ChangeList
from django.core.exceptions import ValidationError
from django.db.models import (
    Case,
    CharField,
    Exists,
    F,
    FloatField,
    IntegerField,
    Max,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Cast, Coalesce, Concat, Round
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path, reverse
from django.utils.html import format_html
from django.utils.translation import gettext as _, ngettext
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from rest_framework.reverse import reverse as drf_reverse

from campaign.models import Campaign, Order
from inventory.models import Category, Product, Variation
from lib.filters import MultiSelectFilter
from lib.models import StringAgg

from .forms import SearchForm
from .models import (
    EmployeeOrderProduct,
    LogisticsCenterInboundReceiptLine,
    POOrder,
    PurchaseOrder,
    PurchaseOrderProduct,
    RegularOrderSummary,
    SupplierOrderSummary,
    VoucherOrderSummary,
)
from .tasks import (
    export_order_summaries_as_xlsx_task,
    send_purchase_orders_to_supplier,
)
from .utils import (
    exclude_taxes,
    get_variation_type,
    get_variations_string,
    snake_to_title,
)


def custom_titled_filter(title):
    class Wrapper(admin.FieldListFilter):
        def __new__(cls, *args, **kwargs):
            instance = admin.FieldListFilter.create(*args, **kwargs)
            instance.title = title
            return instance

    return Wrapper


class POFilter(admin.SimpleListFilter):
    title = 'PO Number'
    parameter_name = 'id'
    chunk_size = 10
    _range = []

    def lookups(self, request, model_admin):
        data_list = POOrder.objects.order_by('id').values_list('id')
        self._range = [
            f'{str(data_list[i : i + self.chunk_size][0][0])}-{str(data_list[i : i + self.chunk_size][-1][0])}'  # noqa: E501
            for i in range(0, len(data_list), self.chunk_size)
        ]
        return [(el, el) for el in self._range]

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        min = self.value().split('-')[0]
        max = self.value().split('-')[1]
        return queryset.filter(id__lte=int(max), id__gte=int(min))


@admin.register(POOrder)
class POOrdersAdmin(admin.ModelAdmin):
    add_form_template = 'logistics/products_order.html'
    change_form_template = 'logistics/products_order.html'
    change_list_template = 'logistics/po_change_list.html'
    list_display = (
        'po_number',
        'supplier',
        'status',
        'created_at',
        'total_cost',
        'logistics_center_status',
        'total_arrived',
    )
    list_filter = (
        ('supplier__name', custom_titled_filter('Supplier Name')),
        (
            'purchaseorderproduct__product_id__brand__name',
            custom_titled_filter('Brand Name'),
        ),
        (POFilter),
        ('status', custom_titled_filter('Status')),
        ('created_at', admin.DateFieldListFilter),
    )
    actions = ['export_as_excel', 'send_again', 'quick_approve', 'cancel_po']
    search_fields = ['id', 'supplier__name', 'status']

    def get_queryset(self, request):
        queryset = super().get_queryset(request)

        # Get total cost from PurchaseOrderProducts
        total_cost_subquery = Subquery(
            PurchaseOrderProduct.objects.filter(purchase_order=OuterRef('id'))
            .annotate(
                base_voucher_value=Case(
                    When(
                        Q(product_id__product_kind='MONEY')
                        & Q(order_products__voucher_val__isnull=False),
                        then=F('order_products__voucher_val'),
                    ),
                    default=Value(None),
                    output_field=FloatField(),
                ),
                base_ordered_quantity=F('quantity_ordered'),
                cost_price=Case(
                    When(
                        product_id__product_kind='MONEY',
                        then=F('voucher_value')
                        * (1 - (F('product_id__supplier_discount_rate') / 100)),
                    ),
                    default=F('product_id__cost_price')
                    / ((100 + settings.TAX_PERCENT) / 100),
                    output_field=IntegerField(),
                ),
            )
            .values('purchase_order')
            .annotate(
                total_cost_db=Round(
                    Sum(F('base_ordered_quantity') * F('cost_price')), 2
                ),
            )
            .values('total_cost_db')
        )

        # Annotate total products arrived count
        queryset = queryset.annotate(
            total_arrived=Sum(
                Case(
                    When(
                        purchaseorderproduct__logisticscenterinboundreceiptline__purchase_order_product__purchase_order=F(
                            'id'
                        ),
                        then=F(
                            'purchaseorderproduct__logisticscenterinboundreceiptline__quantity_received'
                        ),
                    ),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            )
        )

        return queryset.annotate(total_cost_db=total_cost_subquery)

    def total_cost(self, obj):
        return obj.total_cost_db

    total_cost.admin_order_field = 'total_cost_db'

    def total_arrived(self, obj):
        return obj.total_arrived or 0 if obj.status == 'APPROVED' else '-'

    total_arrived.admin_order_field = 'total_arrived'
    total_arrived.short_description = 'Total Arrived'

    def changelist_view(self, request):
        return super().changelist_view(
            request,
            extra_context={
                'title': 'Reports',
                'subtitle': 'PO Orders',
                'has_add_permission': False,
            },
        )

    def changeform_view(self, request, object_id=None, form_url='', extra_context=None):
        """
        Customize the context passed to the change form template.
        """
        if request.GET.get('export'):
            return self.export_as_excel(
                request, self.get_queryset(request).filter(id=object_id)
            )
        if extra_context is None:
            extra_context = {}
        instance: POOrder = self.get_object(request, object_id)
        po_product_variation_mapping = {}
        po_products = []
        if instance:
            po_products = instance.products.annotate(
                po_product_id=Coalesce(
                    F('product_id__bundles__bundle__id'), F('product_id__id')
                ),
                po_quantity=Case(
                    When(
                        product_id__bundles__bundle__id__isnull=False,
                        then=Coalesce(
                            F('order_product__quantity'),
                            Value(0),
                        ),
                    ),
                    default=F('quantity_ordered'),
                    output_field=IntegerField(),
                ),
            )
            po_products = po_products.values('po_product_id').annotate(
                po_quantity=Max('po_quantity'),
                po_id=Max('id'),
                po_product_voucher_value=Max('voucher_value'),
            )
            products = [*instance.products.values()]
            for product in products:
                variation_string = get_variations_string(product.get('variations'))
                product['variation_string'] = variation_string
                po_product_variation_mapping[str(product.get('id'))] = variation_string

        extra_context.update(
            {
                'product_instance': instance,
                'po_products': po_products,
                'variations': po_product_variation_mapping,
                'tax_percent': settings.TAX_PERCENT,
            }
        )

        initial_po_data = []
        if request.GET.get('i'):
            initial_data_key = f'poi_{request.GET.get("i")}'

            if initial_data_key in request.session:
                initial_po_data = request.session[initial_data_key]

        extra_context.update({'initial_po_data': json.dumps(initial_po_data)})

        if instance and instance.status == POOrder.Status.APPROVED.name:
            self.change_form_template = 'logistics/change_products_order.html'
        else:
            self.change_form_template = 'logistics/products_order.html'
        return super().changeform_view(
            request, object_id, form_url, extra_context=extra_context
        )

    def get_urls(self):
        """
        Adding a url to show orders summary.
        """
        urls = super().get_urls()
        my_urls = [
            path('orders-summary/', self.get_orders_summary),
        ]
        return my_urls + urls

    def get_order_workbook(self, order_id):
        workbook = Workbook()
        po_order = POOrder.objects.filter(id=order_id).first()
        if not po_order:
            return workbook
        po_order_supplier = getattr(po_order, 'supplier', None)
        workbook.remove(workbook.active)

        # --- PO Sheet ---
        po_worksheet = workbook.create_sheet('PO')
        po_products = []
        po_products_total_price = 0
        for order_product in po_order.products.all():
            product = getattr(order_product, 'product_id', None)
            if not product:
                continue
            brand = getattr(product, 'brand', None)
            supplier = getattr(product, 'supplier', None)
            is_money_product = getattr(product, 'product_kind', None) == getattr(
                getattr(Product, 'ProductKindEnum', None), 'MONEY', None
            )
            cost_price, total_price, voucher_value = self.get_prices(
                order_product, product
            )
            supplier_discount = getattr(product, 'supplier_discount_rate', None)
            supplier_discount_str = (
                f'{supplier_discount}%' if supplier_discount is not None else '0%'
            )
            variation_string = get_variations_string(order_product.variations)
            sku_with_variation = f'{getattr(product, "sku", "")}{variation_string}'
            if is_money_product:
                po_products_dic = {
                    'SKU': sku_with_variation,
                    'Product Name': getattr(product, 'name_he', ''),
                    'Voucher Type': getattr(product, 'voucher_type', ''),
                    'Voucher Value': voucher_value,
                    'Supplier Discount': supplier_discount_str,
                    'Cost': cost_price,
                    'Quantity': getattr(order_product, 'quantity_ordered', 0),
                    'Total cost': total_price,
                    'Supplier': getattr(supplier, 'name_he', ''),
                }
            else:
                po_products_dic = {
                    'id': getattr(product, 'pk', ''),
                    'name': getattr(product, 'name_he', ''),
                    'category': getattr(
                        getattr(
                            product.categories.all().first()
                            if hasattr(product, 'categories')
                            else None,
                            'name',
                            '',
                        ),
                        'strip',
                        lambda: '',
                    )(),
                    'brand': getattr(brand, 'name_he', ''),
                    'quantity': getattr(order_product, 'quantity_ordered', 0),
                    'quantity_received': 0,
                    'sku': sku_with_variation,
                    'barcode': getattr(product, 'reference', ''),
                    'cost_price': cost_price,
                    'status': getattr(po_order, 'status', ''),
                    'supplier': getattr(supplier, 'name_he', ''),
                    'total_price': total_price,
                }
            po_products.append(po_products_dic)
            po_products_total_price += total_price
        if po_products:
            # Add Description and Total Price to the columns
            po_products_columns = snake_to_title(list(po_products[0].keys())) + [
                'Description',
                'Total Price',
            ]
            po_worksheet.append(po_products_columns)

            # Add Description and Total Price to each product row
            for product in po_products:
                row_values = list(product.values())
                # Add Description and Total Price to the row
                row_values.extend(
                    [getattr(po_order, 'notes', ''), po_products_total_price]
                )
                po_worksheet.append(row_values)
        else:
            po_worksheet.append(['No products found for this PO'])

        # --- Orders Sheet ---
        orders_worksheet = workbook.create_sheet('Orders')
        order_products = []
        for po_product in po_order.products.all():
            for order_product in po_product.order_products.all():
                order = getattr(order_product, 'order_id', None)
                product = getattr(
                    getattr(order_product, 'product_id', None), 'product_id', None
                )
                if not order or not product:
                    continue
                is_money_product = getattr(product, 'product_kind', None) == getattr(
                    getattr(Product, 'ProductKindEnum', None), 'MONEY', None
                )
                if is_money_product:
                    order_products.append(
                        {
                            'Organization name': getattr(
                                getattr(
                                    getattr(order, 'campaign_employee_id', None),
                                    'campaign',
                                    None,
                                ),
                                'organization',
                                None,
                            )
                            and getattr(
                                getattr(
                                    getattr(order, 'campaign_employee_id', None),
                                    'campaign',
                                    None,
                                ).organization,
                                'name_he',
                                '',
                            ),
                            'Group budget': getattr(
                                getattr(
                                    getattr(order_product, 'product_id', None),
                                    'employee_group_campaign_id',
                                    None,
                                ),
                                'budget_per_employee',
                                '',
                            ),
                            'Employee name': getattr(
                                getattr(
                                    getattr(order, 'campaign_employee_id', None),
                                    'employee',
                                    None,
                                ),
                                'full_name_he',
                                '',
                            ),
                            'Phone': getattr(order, 'phone_number', ''),
                            'Voucher value': getattr(order_product, 'voucher_val', ''),
                            'Voucher type': getattr(product, 'voucher_type', ''),
                            'Quantity': getattr(order_product, 'quantity', 0),
                            'Supplier name': getattr(
                                getattr(product, 'supplier', None), 'name_he', ''
                            ),
                            'Product name': getattr(product, 'name_he', ''),
                            'PO number': getattr(po_order, 'id', ''),
                            'Code': None,
                        }
                    )
                else:
                    employee = getattr(order, 'campaign_employee_id', None)
                    employee_obj = (
                        getattr(employee, 'employee', None) if employee else None
                    )
                    full_name = (
                        getattr(employee_obj, 'full_name', '') if employee_obj else ''
                    )
                    order_products.append(
                        {
                            'product_name': getattr(product, 'name_he', ''),
                            'sku': getattr(product, 'sku', ''),
                            'quantity': getattr(order_product, 'quantity', 0),
                            'cost_price': round(
                                getattr(product, 'cost_price', 0)
                                / ((100 + getattr(settings, 'TAX_PERCENT', 0)) / 100),
                                2,
                            )
                            if getattr(product, 'cost_price', None) is not None
                            else 0,
                            'order_id': getattr(order, 'pk', ''),
                            'organization_name': getattr(
                                getattr(
                                    getattr(order, 'campaign_employee_id', None),
                                    'campaign',
                                    None,
                                ),
                                'organization',
                                None,
                            )
                            and getattr(
                                getattr(
                                    getattr(order, 'campaign_employee_id', None),
                                    'campaign',
                                    None,
                                ).organization,
                                'name_he',
                                '',
                            ),
                            'employee_name': f'{full_name} - NKS',
                            'phone_number': getattr(order, 'phone_number', ''),
                            'additional_phone_number': getattr(
                                order, 'additional_phone_number', ''
                            ),
                            'delivery_street': getattr(order, 'delivery_street', ''),
                            'delivery_street_number': getattr(
                                order, 'delivery_street_number', ''
                            ),
                            'delivery_apartment_number': getattr(
                                order, 'delivery_apartment_number', ''
                            ),
                            'delivery_city': getattr(order, 'delivery_city', ''),
                            'delivery_additional_details': getattr(
                                order, 'delivery_additional_details', ''
                            ),
                            'supplier_name': getattr(po_order_supplier, 'name', ''),
                        }
                    )
        if order_products:
            order_products_columns = snake_to_title(list(order_products[0].keys()))
            orders_worksheet.append(order_products_columns)
            for product in order_products:
                orders_worksheet.append(list(product.values()))
        else:
            orders_worksheet.append(['No order products found for this PO'])
        return workbook

    def get_orders_summary(self, request, **kwargs):
        """
        Showing the orders summary with filtering and sorting.
        """
        search_form = SearchForm(request.GET)
        model_name = self.model._meta.model_name
        back = f'admin:{self.model._meta.app_label}_{model_name}_changelist'

        # Get filter parameters from GET request
        supplier = request.GET.get('supplier', '')
        brand = request.GET.get('brand', '')

        # Build the queryset with filters
        qs = (
            PurchaseOrderProduct.objects.values(
                'product_id__id',
                'product_id__supplier__name',
                'product_id__brand__name',
                'product_id__sku',
                'product_id__reference',
                'product_id__cost_price',
            )
            .annotate(
                total_ordered=Sum('quantity_ordered'),
            )
            .annotate(
                total_in_dc=Coalesce(
                    Sum(
                        Subquery(
                            LogisticsCenterInboundReceiptLine.objects.filter(
                                purchase_order_product=OuterRef('pk')
                            )
                            .values('purchase_order_product')
                            .order_by('purchase_order_product')
                            .annotate(total_in_dc=Sum('quantity_received'))
                            .values('total_in_dc')[:1],
                        )
                    ),
                    0,
                )
            )
            .annotate(
                total_in_transit=Sum('quantity_sent_to_logistics_center')
                - F('total_in_dc'),
                difference_to_order=F('total_ordered')
                - F('total_in_transit')
                - F('total_in_dc'),
            )
        )

        if search_form.is_valid():
            query = search_form.cleaned_data['query']
            if query:
                qs = qs.filter(
                    Q(product_id__supplier__name__icontains=query)
                    | Q(product_id__brand__name__icontains=query)
                    | Q(product_id__sku__icontains=query)
                    | Q(product_id__reference__icontains=query)
                )

        if supplier:
            qs = qs.filter(product_id__supplier__name__icontains=supplier)
        if brand:
            qs = qs.filter(product_id__brand__name__icontains=brand)

        return render(
            request,
            'logistics/po_summary.html',
            {
                'summary': list(qs),
                'back': back,
                'search_form': search_form,
                'filter': {
                    'supplier': supplier,
                    'brand': brand,
                },
            },
        )

    def get_voucher_value(self, po_product):
        """
        Calculate the voucher value for each product in the PurchaseOrder.
        """
        if po_product and po_product.voucher_value:
            return po_product.voucher_value
        return 0

    def get_prices(self, order_product, product):
        """
        Calculate the cost price, total price, and voucher value for an order product.

        Args:
            order_product: Object containing details of the ordered product
            product: Object representing the product with its pricing and kind.

        Returns:
            Tuple (cost_price, total_price, voucher_value):
                - cost_price (float)
                - total_price (float)
                - voucher_value (float)
        """

        is_money_product = product.product_kind == Product.ProductKindEnum.MONEY.name
        voucher_value = 0

        if is_money_product:
            voucher_value = self.get_voucher_value(order_product)
            supplier_discount = product.supplier_discount_rate / 100 or 0
            cost_price = int(
                (round(voucher_value * (1 - supplier_discount), 2))
                if supplier_discount
                else (voucher_value)
            )
        else:
            cost_price = int(
                round(product.cost_price / ((100 + settings.TAX_PERCENT) / 100), 2)
            )

        total_price = cost_price * order_product.quantity_ordered

        return cost_price, total_price, voucher_value

    def create_po_row(self, order, product, merge=False):
        if merge:
            record = [
                '',
                product.product_name,
                '',
                '',
                product.quantity_ordered,
                product.quantity_received,
                product.sku,
                product.barcode,
                product.cost_price,
                '',
                '',
                exclude_taxes(product.total_cost),
                product.description,
            ]
            return record
        record = [
            order.id,
            product.product_name,
            product.category,
            product.brand,
            product.quantity_ordered,
            product.quantity_received,
            product.sku,
            product.barcode,
            product.cost_price,
            order.status,
            order.supplier.name,
            exclude_taxes(product.total_cost),
            product.description,
        ]
        return record

    def create_orders_row(self, product, merge=False):
        if merge:
            record = [
                product.product_name,
                product.sku,
                product.quantity_ordered,
                product.cost_price,
                product.order_id,
            ]
            return record
        record = [
            product.product_name,
            product.sku,
            product.quantity_ordered,
            product.cost_price,
            product.order_id,
            product.organization,
            product.employee_name,
            product.phone_number,
            product.additional_phone_number,
            product.delivery_street,
            product.delivery_street_number,
            product.delivery_apartment_number,
            product.delivery_city,
            product.delivery_additional_details,
            product.supplier_name,
        ]
        return record

    def export_as_excel(self, request, queryset):
        """
        Export all queried orders as separate xlsx files, and if there are
        multiple files, download them all in a zip folder. If there's only one
        file, download it directly.
        """
        po_field_names = [
            'Id',
            'Name',
            'Category',
            'Brand',
            'Quantity',
            'Quantity Received',
            'Sku',
            'Barcode',
            'Cost Price',
            'Status',
            'Supplier',
            'Total Price',
            'Description',
        ]

        orders_field_names = [
            'ProductName',
            'SKU',
            'Quantity',
            'Cost Price',
            'Order Id',
            'Organization',
            'Employee Name',
            'Phone Number',
            'Additional Phone Number',
            'Delivery Street',
            'Delivery Street Number',
            'Delivery Apartment Number',
            'Delivery City',
            'Delivery Additional Details',
            'Supplier Name',
        ]

        queryset = queryset.prefetch_related(
            Prefetch(
                'purchaseorderproduct_set',
                queryset=PurchaseOrderProduct.objects.annotate(
                    supplier_name=F('purchase_order__supplier__name'),
                    product_name=F('product_id__name_he'),
                    brand=F('order_product__product_id__product_id__brand__name_he'),
                    category=Subquery(
                        Category.objects.filter(
                            product__id=OuterRef(
                                'order_product__product_id__product_id__id'
                            )
                        )
                        .values('name_he')
                        .annotate(categories=StringAgg('name_he'))
                        .values('categories')
                    ),
                    quantity_received=Coalesce(
                        F('logisticscenterinboundreceiptline__quantity_received'),
                        Value(0),
                    ),
                    sku=F('product_id__sku'),
                    barcode=F('product_id__reference'),
                    cost_price=F('product_id__cost_price'),
                    description=F('product_id__description_he'),
                    employee_name=Concat(
                        F(
                            'order_product__order_id__campaign_employee_id__employee__first_name'
                        ),
                        Value(' '),
                        F(
                            'order_product__order_id__campaign_employee_id__employee__last_name'
                        ),
                    ),
                    phone_number=F('order_product__order_id__phone_number'),
                    additional_phone_number=F(
                        'order_product__order_id__additional_phone_number'
                    ),
                    delivery_city=F('order_product__order_id__delivery_city'),
                    delivery_street=F('order_product__order_id__delivery_street'),
                    delivery_street_number=F(
                        'order_product__order_id__delivery_street_number'
                    ),
                    delivery_apartment_number=F(
                        'order_product__order_id__delivery_apartment_number'
                    ),
                    delivery_additional_details=F(
                        'order_product__order_id__delivery_additional_details'
                    ),
                    order_quantity=F('order_product__quantity'),
                    organization=F(
                        'order_product__order_id__campaign_employee_id__campaign__organization__name_he'
                    ),
                    order_id=F('order_product__order_id__pk'),
                    order_product_sku=F('order_product__product_id__product_id__sku'),
                    order_product_kind=F(
                        'order_product__product_id__product_id__product_kind'
                    ),
                ).order_by('order_product__product_id__product_id__sku'),
                to_attr='po_products',
            )
        )

        if queryset.count() == 1:
            # Handle single record case
            order = queryset.first()
            workbook = Workbook()
            workbook.remove(workbook.active)
            po_worksheet = workbook.create_sheet('PO')
            orders_worksheet = workbook.create_sheet('Orders')
            po_worksheet.append(po_field_names)
            orders_worksheet.append(orders_field_names)
            bundle_products = {}

            for idx, product in enumerate(order.po_products):
                merge = False
                if product.order_product_kind == Product.ProductKindEnum.BUNDLE.name:
                    if product.order_product_sku not in bundle_products.keys():
                        bundle_products[product.order_product_sku] = {
                            'start': idx + 2,
                            'end': idx + 2,
                        }
                    else:
                        bundle_products[product.order_product_sku]['end'] = (
                            bundle_products[product.order_product_sku]['end'] + 1
                        )
                        merge = True
                record = self.create_po_row(order, product, merge)
                po_worksheet.append(record)
                record = self.create_orders_row(product, merge)
                orders_worksheet.append(record)

            for merge_values in bundle_products.values():
                if merge_values.get('end') != merge_values.get('start'):
                    for _column in range(5, 16):
                        orders_worksheet.merge_cells(
                            start_row=merge_values.get('start'),
                            start_column=_column,
                            end_row=merge_values.get('end'),
                            end_column=_column,
                        )
                        cell = orders_worksheet[
                            f'{get_column_letter(_column)}{merge_values.get("start")}'
                        ]
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                    for _column in [1, 3, 4, 10, 11]:
                        po_worksheet.merge_cells(
                            start_row=merge_values.get('start'),
                            start_column=_column,
                            end_row=merge_values.get('end'),
                            end_column=_column,
                        )
                        cell = po_worksheet[
                            f'{get_column_letter(_column)}{merge_values.get("start")}'
                        ]
                        cell.alignment = Alignment(vertical='center', wrap_text=True)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (
                'attachment; filename=order_{}.xlsx'.format(order.po_number)
            )
            workbook.save(response)
            return response
        else:
            # Handle multiple records case
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                for order in queryset:
                    workbook = Workbook()
                    workbook.remove(workbook.active)
                    po_worksheet = workbook.create_sheet('PO')
                    orders_worksheet = workbook.create_sheet('Orders')
                    po_worksheet.append(po_field_names)
                    orders_worksheet.append(orders_field_names)
                    bundle_products = {}

                    for idx, product in enumerate(order.po_products):
                        merge = False
                        if (
                            product.order_product_kind
                            == Product.ProductKindEnum.BUNDLE.name
                        ):
                            if product.order_product_sku not in bundle_products.keys():
                                bundle_products[product.order_product_sku] = {
                                    'start': idx + 2,
                                    'end': idx + 2,
                                }
                            else:
                                bundle_products[product.order_product_sku]['end'] = (
                                    bundle_products[product.order_product_sku]['end']
                                    + 1
                                )
                                merge = True
                        record = self.create_po_row(order, product, merge)
                        po_worksheet.append(record)
                        record = self.create_orders_row(product, merge)
                        orders_worksheet.append(record)
                    for merge_values in bundle_products.values():
                        if merge_values.get('end') != merge_values.get('start'):
                            for _column in range(5, 16):
                                orders_worksheet.merge_cells(
                                    start_row=merge_values.get('start'),
                                    start_column=_column,
                                    end_row=merge_values.get('end'),
                                    end_column=_column,
                                )
                                cell = orders_worksheet[
                                    f'{get_column_letter(_column)}{merge_values.get("start")}'
                                ]
                                cell.alignment = Alignment(
                                    vertical='center', wrap_text=True
                                )
                            for _column in [1, 3, 4, 10, 11]:
                                po_worksheet.merge_cells(
                                    start_row=merge_values.get('start'),
                                    start_column=_column,
                                    end_row=merge_values.get('end'),
                                    end_column=_column,
                                )
                                cell = po_worksheet[
                                    f'{get_column_letter(_column)}{merge_values.get("start")}'
                                ]
                                cell.alignment = Alignment(
                                    vertical='center', wrap_text=True
                                )
                    # Save workbook to a BytesIO object and add it to the zip file
                    file_buffer = io.BytesIO()
                    workbook.save(file_buffer)
                    file_buffer.seek(0)
                    zip_file.writestr(
                        'order_{}.xlsx'.format(order.po_number), file_buffer.read()
                    )

            response = HttpResponse(
                zip_buffer.getvalue(), content_type='application/zip'
            )
            response['Content-Disposition'] = 'attachment; filename=orders.zip'
            return response

    def send_again(self, request, queryset):
        send_purchase_orders_to_supplier.apply_async(
            (list(queryset.values_list('id', flat=True)),)
        )
        self.message_user(
            request,
            ngettext(
                'Sent email to the supplier',
                'Sent emails to all the suppliers',
                queryset.count(),
            ),
            messages.SUCCESS,
        )

    def quick_approve(self, request, queryset):
        approved_count = 0
        errors = []

        # update status with calls to save for signals to be invoked
        for po in queryset.all():
            try:
                po.approve()
                approved_count += 1
            except ValidationError as ex:
                errors.append(ex.message)

        if approved_count > 0:
            self.message_user(
                request,
                ngettext(
                    'Purchase order is approved',
                    'Purchase orders are approved',
                    approved_count,
                ),
                messages.SUCCESS,
            )
        if len(errors) > 0:
            self.message_user(
                request,
                ', '.join(errors),
                messages.ERROR,
            )

    def cancel_po(self, request, queryset):
        queryset.update(status=POOrder.Status.CANCELLED.name)
        self.message_user(
            request,
            ngettext(
                'Purchase order is cancelled',
                'Purchase orders are cancelled',
                queryset.count(),
            ),
            messages.SUCCESS,
        )


class GuaranteedDeterministicChangeList(ChangeList):
    """
    This is a ChangeList descendant which removes the deterministic order logic
    since we want to display aggregated data in change lists, which the
    original logic deems non-deterministcally-ordered and then adds ordering
    (and a groups) by the main object pk, which then ruins our aggregation.
    Do not use this for normal admins!
    """

    def _get_deterministic_ordering(self, ordering):
        return list(ordering)


class CampaignStatusFilter(MultiSelectFilter):
    custom_title = 'campaign status'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Campaign.CampaignStatusEnum)]


class SupplierFilter(MultiSelectFilter):
    custom_title = 'supplier'


class BrandFilter(MultiSelectFilter):
    custom_title = 'brand'


class ProductTypeFilter(MultiSelectFilter):
    custom_title = 'product type'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Product.ProductTypeEnum)]


class ProductKindFilter(MultiSelectFilter):
    custom_title = 'product kind'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Product.ProductKindEnum)]


class ProductVoucherTypeFilter(MultiSelectFilter):
    custom_title = 'Voucher Type'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Product.VoucherTypeEnum)]


class OrganizationFilter(MultiSelectFilter):
    custom_title = 'organization'


class TagsFilter(MultiSelectFilter):
    custom_title = 'tags'


class CampaignTypeFilter(MultiSelectFilter):
    custom_title = 'campaign type'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Campaign.CampaignTypeEnum)]


class ExcludeColumnsFilter(admin.SimpleListFilter):
    """
    Admin filter to exclude certain columns
    """

    title = 'Exclude Columns'
    parameter_name = 'exclude_columns'

    def lookups(self, request, model_admin):
        list_display = model_admin.get_list_display(request)
        return [(column, column) for column in list_display]

    def queryset(self, request, queryset):
        return queryset

    def choices(self, changelist):
        return []

    def has_output(self):
        return False


class IncludeColumnsFilter(admin.SimpleListFilter):
    """
    Admin filter to Include certain columns
    """

    title = 'Include Columns'
    parameter_name = 'include_columns'

    def lookups(self, request, model_admin):
        list_display = model_admin.get_list_display(request)
        return [(column, column) for column in list_display]

    def queryset(self, request, queryset):
        return queryset

    def choices(self, changelist):
        return []

    def has_output(self):
        return False


class ColumnsOrderFilter(admin.SimpleListFilter):
    """
    Admin filter to Order certain columns
    """

    title = 'Columns Order'
    parameter_name = 'columns_order'

    def lookups(self, request, model_admin):
        list_display = model_admin.get_list_display(request)
        return [(column, column) for column in list_display]

    def queryset(self, request, queryset):
        return queryset

    def choices(self, changelist):
        return []

    def has_output(self):
        return False


class StickyEAdminMixin:
    """
    Mixin for Django ModelAdmin to preserve a custom 'e' GET parameter
    across all filter and sort links in the admin changelist.

    This works by monkey-patching the ChangeList instance returned by
    get_changelist_instance(): wrapping its get_query_string method
    so that, whenever the admin builds a URL for filters or sorting,
    it re-appends the original request’s 'e' parameter if present.
    """

    def get_changelist_instance(self, request):
        changelist = super().get_changelist_instance(request)
        original_gqs = changelist.get_query_string

        def get_query_string(new_params=None, remove=None):
            qs = original_gqs(new_params, remove)
            e = request.GET.get('e')
            if e == 'regular':
                if e and 'e=' not in qs:
                    sep = '&' if '?' in qs else '?'
                    return f'{qs}{sep}e={e}'
                return qs
            return qs

        changelist.get_query_string = get_query_string
        return changelist


@admin.register(EmployeeOrderProduct)
class OrderSummaryAdmin(StickyEAdminMixin, admin.ModelAdmin):
    class Media:
        css = {'all': ('admin/css/freeze_order_summary_header.css',)}
        js = ('admin/js/freeze_order_summary_header.js',)

    list_display_links = None
    ordering = ['product_id__product_id__sku']
    list_filter = (
        (
            'product_id__employee_group_campaign_id__campaign__status',
            CampaignStatusFilter,
        ),
        ('product_id__product_id__supplier__name', SupplierFilter),
        ('product_id__product_id__brand__name', BrandFilter),
        ('product_id__product_id__product_type', ProductTypeFilter),
        ('product_id__product_id__product_kind', ProductKindFilter),
        ('product_id__product_id__voucher_type', ProductVoucherTypeFilter),
        (
            'product_id__employee_group_campaign_id__campaign__organization__name',
            OrganizationFilter,
        ),
        ('product_id__employee_group_campaign_id__campaign__tags__name', TagsFilter),
        (
            'product_id__employee_group_campaign_id__campaign__campaign_type',
            CampaignTypeFilter,
        ),
        (ExcludeColumnsFilter),
        (IncludeColumnsFilter),
        (ColumnsOrderFilter),
    )
    search_fields = (
        'product_sku',
        'product_supplier',
        'product_brand',
    )
    actions = ('create_po', 'export_as_xlsx')

    change_list_template = 'logistics/order_summary_change_list.html'

    def get_list_display(self, request):
        path_only = request.path
        if 'regularordersummary' in path_only or 'supplierordersummary' in path_only:
            return [
                'product_supplier',
                'total_ordered',
                'sent_to_approve',
                'in_transit_stock',
                'difference_to_order',
            ]

        if 'voucherordersummary' in path_only:
            return [
                'product_supplier',
                'total_ordered',
                'voucher_type',
                'sent_to_approve',
                'in_transit_stock',
                'difference_to_order',
            ]
        list_display = [
            'product_supplier',
            'product_brand',
            'product_sku',
            'product_kind',
        ]

        variations = self.get_queryset(request=request).values('variations')
        extra_keys = self.get_extra_keys(variations=variations)

        list_display.extend(extra_keys)
        list_display.extend(
            [
                'product_name',
                'product_reference',
                'voucher_value',
                'total_ordered',
                'product_cost_price',
                'product_quantity',
                'sent_to_approve',
                'in_transit_stock',
                'dc_stock',
                'product_snapshot_stock',
                'difference_to_order',
                'product_snapshot_stock_date_time',
            ]
        )

        for key in extra_keys:
            self.create_dynamic_column(key)

        exclude_columns = request.GET.get('exclude_columns')
        if exclude_columns:
            exclude_columns = exclude_columns.split(',')
            for column in exclude_columns:
                if column in list_display:
                    list_display.remove(column)
                if column == 'variations':
                    for key in extra_keys:
                        list_display.remove(key)

        include_columns = request.GET.get('include_columns')
        if include_columns:
            include_columns = include_columns.split(',')
            for column in include_columns:
                if column not in list_display:
                    list_display.append(column)
                    self.create_dynamic_column(column)

        columns_order = request.GET.get('columns_order')
        if columns_order:
            list_display = []
            for column in columns_order.split(','):
                if column == 'variations':
                    list_display.extend(extra_keys)
                else:
                    list_display.append(column)

        return list_display

    def get_extra_keys(self, variations):
        unique_keys = set()
        for extras in variations:
            if isinstance(extras, dict):
                variation_data = extras.values()
                for variation in variation_data:
                    if isinstance(variation, dict):
                        unique_keys.update(variation.keys())
        unique_keys = [key.replace(' ', '_') for key in unique_keys]
        return list(unique_keys)

    def create_dynamic_column(self, key):
        def dynamic_column(admin_self, obj):
            merged = obj.get('merged_variations', {})
            return merged.get(key, '')

        dynamic_column.short_description = key.replace('_', ' ')
        setattr(OrderSummaryAdmin, key, dynamic_column)

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        path_only = request.path
        qs = super().get_queryset(request)
        qs = qs.filter(
            order_id__status__in=(
                Order.OrderStatusEnum.PENDING.name,
                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
            ),
        )

        # Define subqueries for sent_to_dc, quantity_ordered, and dc_stock.
        sent_to_dc_subquery = Subquery(
            PurchaseOrderProduct.objects.filter(
                product_id_id=OuterRef('product_id__product_id_id'),
            )
            .filter(
                Q(voucher_value=OuterRef('base_voucher_value'))
                | Q(voucher_value=None)
                | Q(voucher_value=0)
            )
            .values('product_id_id')
            .annotate(sent_to_dc=Sum(F('quantity_sent_to_logistics_center')))
            .values('sent_to_dc')[:1]
        )

        quantity_ordered_subquery = Subquery(
            PurchaseOrderProduct.objects.filter(
                purchase_order__status=PurchaseOrder.Status.SENT_TO_SUPPLIER.name,
                product_id_id=OuterRef('product_id__product_id_id'),
            )
            .filter(
                Q(voucher_value=OuterRef('base_voucher_value'))
                | Q(voucher_value=None)
                | Q(voucher_value=0)
            )
            .values('product_id_id', 'voucher_value')
            .annotate(quantity_ordered=Sum(F('quantity_ordered')))
            .values('quantity_ordered')[:1]
        )

        dc_stock_subquery = Subquery(
            PurchaseOrderProduct.objects.filter(
                product_id_id=OuterRef('product_id__product_id_id')
            )
            .filter(
                Q(voucher_value=OuterRef('base_voucher_value'))
                | Q(voucher_value=None)
                | Q(voucher_value=0)
            )
            .values('product_id_id')
            .annotate(
                dc_stock=Sum(F('logisticscenterinboundreceiptline__quantity_received'))
            )
            .values('dc_stock')[:1]
        )

        # Annotate base fields.
        base_qs = qs.annotate(
            product_supplier=F('product_id__product_id__supplier__name'),
            product_sku=F('product_id__product_id__sku'),
            product_kind=F('product_id__product_id__product_kind'),
            product_name=F('product_id__product_id__name_he'),
            base_ordered_quantity=F('quantity'),
            base_voucher_value=Case(
                When(
                    Q(product_kind='MONEY') & Q(voucher_val__isnull=False),
                    then=F('voucher_val'),
                ),
                default=Value(None),
                output_field=FloatField(),
            ),
            voucher_type=F('product_id__product_id__voucher_type'),
        ).annotate(
            sent_to_dc=Coalesce(sent_to_dc_subquery, 0),
            dc_stock=Coalesce(dc_stock_subquery, 0),
        )
        self.unmerged_qs = base_qs

        # Group by based on the request.
        group_by = [
            'product_supplier',
            'product_sku',
            'base_voucher_value',
        ]
        if 'regularordersummary' in path_only or 'supplierordersummary' in path_only:
            group_by = ['product_supplier']

        if 'voucherordersummary' in path_only:
            group_by = ['product_supplier', 'voucher_type']

        if 'regularordersummary' in path_only or request.GET.get('e') == 'regular':
            matching_difference_to_order = (
                base_qs.annotate(
                    sent_to_dc_product=Coalesce(sent_to_dc_subquery, 0),
                )
                .annotate(
                    total_ordered_product=Sum('base_ordered_quantity'),
                    sent_to_approve_product=Coalesce(quantity_ordered_subquery, 0),
                )
                .annotate(
                    difference_to_order=F('total_ordered_product')
                    - F('sent_to_dc_product')
                    - F('sent_to_approve_product'),
                )
                .annotate(
                    voucher_type_norm=Coalesce(F('voucher_type'), Value('voucher_type'))
                )
                .filter(
                    Q(difference_to_order__gt=0)
                    & Q(product_supplier=OuterRef('product_supplier'))
                    & Q(product_sku=OuterRef('product_sku'))
                    & Q(base_voucher_value=None)
                    & Q(
                        voucher_type_norm=Coalesce(
                            OuterRef('product_id__product_id__voucher_type'),
                            Value('voucher_type'),
                        )
                    )
                )
            )
            base_qs = base_qs.filter(Exists(matching_difference_to_order)).filter(
                (
                    Q(
                        product_id__product_id__product_type=Product.ProductTypeEnum.REGULAR.name
                    )
                    | Q(
                        product_id__product_id__product_type=Product.ProductTypeEnum.LARGE_PRODUCT.name
                    )
                )
                & ~Q(
                    product_id__product_id__product_kind=Product.ProductKindEnum.MONEY.name
                )
            )
        elif 'supplierordersummary' in path_only:
            base_qs = base_qs.filter(
                Q(
                    product_id__product_id__product_type=Product.ProductTypeEnum.SENT_BY_SUPPLIER.name
                )
                & ~Q(
                    product_id__product_id__product_kind=Product.ProductKindEnum.MONEY.name
                )
            )
        elif 'voucherordersummary' in path_only:
            base_qs = base_qs.filter(
                product_id__product_id__product_kind=Product.ProductKindEnum.MONEY.name
            )

        grouped_qs = base_qs.values(*group_by)
        if (
            'regularordersummary' in path_only
            or 'supplierordersummary' in path_only
            or 'voucherordersummary' in path_only
        ):
            grouped_qs = (
                grouped_qs.annotate(
                    pk=F('product_supplier'),
                    total_ordered=Sum('base_ordered_quantity'),
                    orders=StringAgg('pk'),
                    sent_to_approve=Coalesce(quantity_ordered_subquery, 0),
                    sum_sent_to_dc=Sum('sent_to_dc'),
                    sum_dc_stock=Sum('dc_stock'),
                )
                .annotate(
                    in_transit_stock=F('sum_sent_to_dc') - F('sum_dc_stock'),
                    difference_to_order=F('total_ordered')
                    - F('sent_to_dc')
                    - F('sent_to_approve'),
                )
                .order_by('-difference_to_order')
            )

        else:
            grouped_qs = (
                grouped_qs.annotate(
                    total_ordered=Sum('base_ordered_quantity'),
                    voucher_value=F('base_voucher_value'),
                    voucher_type=F('voucher_type'),
                    sum_sent_to_dc=Max('sent_to_dc'),
                    sum_dc_stock=Max('dc_stock'),
                    product_brand=F('product_id__product_id__brand__name'),
                    product_reference=F('product_id__product_id__reference'),
                    product_cost_price=F('product_id__product_id__cost_price'),
                    product_quantity=F('product_id__product_id__product_quantity'),
                    product_kind=F('product_id__product_id__product_kind'),
                    product_name=F('product_id__product_id__name_he'),
                    product_pk=F('product_id__product_id_id'),
                    product_snapshot_stock=F(
                        'product_id__product_id__logistics_snapshot_stock_line__quantity'
                    ),
                    product_snapshot_stock_date_time=F(
                        'product_id__product_id__logistics_snapshot_stock_line__stock_snapshot__snapshot_date_time'
                    ),
                    orders=StringAgg('pk'),
                    sent_to_approve=Coalesce(quantity_ordered_subquery, 0),
                )
                .annotate(
                    dc_stock=F('sum_dc_stock'),
                    in_transit_stock=F('sum_sent_to_dc') - F('sum_dc_stock'),
                    difference_to_order=F('total_ordered')
                    - F('sent_to_dc')
                    - F('sent_to_approve'),
                    pk=Concat(
                        Cast('product_pk', CharField()),
                        Value('_'),  # Optional separator
                        Cast('voucher_value', CharField()),
                        Value('_'),
                        Cast('variations', CharField()),
                        output_field=CharField(),
                    ),
                    voucher_value=Coalesce(F('voucher_val'), F('voucher_value')),
                    delivery_price=F('product_id__product_id__delivery_price'),
                )
                .order_by('-difference_to_order')
            )

        final_keys = [
            'product_supplier',
            'product_sku',
            'total_ordered',
            'voucher_type',
            'sum_sent_to_dc',
            'sum_dc_stock',
            'product_brand',
            'product_reference',
            'product_cost_price',
            'product_quantity',
            'product_kind',
            'product_name',
            'sent_to_approve',
            'product_pk',
            'product_snapshot_stock',
            'product_snapshot_stock_date_time',
            'dc_stock',
            'in_transit_stock',
            'difference_to_order',
            'pk',
            'variations',
            'voucher_value',
            'orders',
            'delivery_price',
        ]

        if 'regularordersummary' in path_only or 'supplierordersummary' in path_only:
            final_keys = [
                'product_supplier',
                'total_ordered',
                'sent_to_approve',
                'in_transit_stock',
                'difference_to_order',
                'pk',
            ]

        if 'voucherordersummary' in path_only:
            final_keys = [
                'product_supplier',
                'total_ordered',
                'voucher_type',
                'sent_to_approve',
                'in_transit_stock',
                'difference_to_order',
                'pk',
            ]

        # Preserve extra variation keys as before.
        keys = self.get_extra_keys(variations=grouped_qs.values('variations'))
        grouped_qs = grouped_qs.annotate(**{key: Value('') for key in keys})
        return grouped_qs.values(*final_keys)

    def get_changelist(self, request, **kwargs):
        return GuaranteedDeterministicChangeList

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['active_view'] = request.GET.get('view', 'all')
        extra_context['grouped_data'] = self.get_queryset(request=request)

        page = request.GET.get('e', '')
        # Set dynamic title based on URL path
        if page == 'voucher' or 'voucherordersummary' in request.path:
            extra_context['title'] = 'Order Summaries – Vouchers'
        elif page == 'supplier' or 'supplierordersummary' in request.path:
            extra_context['title'] = 'Order Summaries – Sent by Supplier'
        elif page == 'regular' or 'regularordersummary' in request.path:
            extra_context['title'] = 'Order Summaries – Regular'
        else:
            extra_context['title'] = 'Order Summaries'
        response = super().changelist_view(request, extra_context=extra_context)
        try:
            cl = response.context_data.get('cl')
            if not cl:
                return response

            # Compute extra keys from variations (using the base queryset)
            extra_keys = self.get_extra_keys(
                variations=self.get_queryset(request).values('variations')
            )
            for row in cl.result_list:
                supplier = row.get('product_supplier')
                sku = row.get('product_sku')
                merged_variations = {}
                duplicates = self.unmerged_qs.filter(
                    product_supplier=supplier, product_sku=sku
                ).values_list('variations', flat=True)
                for var in duplicates:
                    if not var:
                        continue
                    for k, v in var.items():
                        merged_variations[k] = v

                # Optionally, append variation text to the SKU for display
                variations_text = []
                for key, value in merged_variations.items():
                    variation_kind = get_variation_type(
                        variation=key
                    )  # your helper function
                    if variation_kind == 'COLOR':
                        variations_text.insert(0, str(value)[:3])
                    elif variation_kind == 'TEXT':
                        variations_text.append(str(value)[:3])

                variations = row['variations']
                if variations:
                    variations_text = []
                    for key, value in variations.items():
                        row[key.replace(' ', '_')] = value
                        variation_kind = get_variation_type(variation=key)
                        if variation_kind == 'COLOR':
                            variations_text.insert(0, str(value[:3]))
                        elif variation_kind == 'TEXT':
                            variations_text.append(str(value[:3]))

                    row['product_sku'] = str(row['product_sku']) + ''.join(
                        variations_text
                    )

                # Ensure every extra dynamic key is present in the row, even if empty.
                for key in extra_keys:
                    if key not in row:
                        row[key] = ''
        except Exception as e:
            print(f'Error merging variations: {e}')
        return response

    @staticmethod
    def parse_variations(order_summary):
        color_variations = []
        text_variations = []
        if order_summary['variations']:
            for variation_type, variation_value in order_summary['variations'].items():
                variation_instance = Variation.objects.filter(
                    Q(site_name_he=variation_type) | Q(site_name_en=variation_type)
                ).first()
                if variation_instance:
                    if (
                        Variation.VariationKindEnum.COLOR.name
                        == variation_instance.variation_kind
                    ):
                        color_variations.append(variation_value)
                    elif (
                        Variation.VariationKindEnum.TEXT.name
                        == variation_instance.variation_kind
                    ):
                        text_variations.append(variation_value)
        return '|'.join([', '.join(color_variations), ', '.join(text_variations)])

    def action_checkbox(self, obj):
        unique_id = str(obj.get('pk', ''))
        # if obj.get('product_kind', '') == "MONEY":
        #     unique_id = str(obj.get('pk', '')) + str(obj.get('voucher_value', ''))
        # if obj.get('product_kind', '') == "VARIATION":
        #     unique_id = str(obj.get('pk', '')) + self.parse_variations(obj)
        #
        # print(unique_id)
        # print(self.parse_variations(obj))
        """
        An override of the ancestor method to use a non-pk field and read it
        from a dictionary and not only as an attribute
        """
        attrs = {
            'class': 'action-select',
            'aria-label': format_html(
                _('Select this object for an action - {}'), unique_id
            ),
        }
        checkbox = forms.CheckboxInput(attrs, lambda value: False)
        return checkbox.render(helpers.ACTION_CHECKBOX_NAME, str(obj['pk']))

    def create_po(self, request, queryset):
        product_kinds = queryset.values_list('product_kind', flat=True)
        if 'MONEY' in product_kinds and len(set(product_kinds)) > 1:
            self.message_user(
                request, 'Please select same kind of products', level=messages.ERROR
            )
            return

        if request.GET.get('q'):
            self.message_user(
                request,
                'Text search cannot be used when creating a purchase order',
                level=messages.ERROR,
            )
            return

        params = {}

        if queryset.first():
            i = int(time.time())
            changelist_filters = request.GET.copy()
            if 'p' in changelist_filters:
                # remove the page queryset parameter if it is set
                changelist_filters.pop('p')
            if 'o' in changelist_filters:
                # remove the order queryset parameter if it is set
                changelist_filters.pop('o')

            params = {
                'supplierName': queryset.first()['product_supplier'],
                'i': i,
                'product_sku': queryset.first()['product_sku'],
                # keep current parameters which include order summaries filters
                'changelistFilters': changelist_filters.urlencode(),
            }

            request.session[f'poi_{i}'] = [
                {
                    'sku': p['product_sku'],
                    'pk': p['pk'],
                    'quantity': max(0, p['difference_to_order']),
                    'variation': (
                        f'{p["product_sku"]}'
                        f'{get_variations_string(p.get("variations"))}'
                    ),
                    'voucher_value': p['voucher_value'],
                    'orders': p['orders'].split(','),
                    'variations_json': p.get('variations'),
                }
                for p in queryset.all()
            ]

        add_page_url = reverse('admin:logistics_poorder_add')
        add_page_url = f'{add_page_url}?{urlencode(params)}'
        return redirect(add_page_url)

    def export_as_xlsx(self, request, queryset):
        # using the inner query since it can be pickled to recreate the
        # queryset (celery is in charge of pickling here), based on docs -
        # https://docs.djangoproject.com/en/5.1/ref/models/querysets/#pickling-querysets
        export_order_summaries_as_xlsx_task.apply_async(
            (
                queryset.query,
                request.user.pk,
                request.user.email,
                drf_reverse('download_export_file_view', request=request),
            )
        )

        self.message_user(
            request,
            (
                'Exporting order summaries, file will be sent to '
                f'"{request.user.email}" when ready'
            ),
            messages.SUCCESS,
        )

    def product_supplier(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_supplier.short_description = 'Supplier'

    def product_brand(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_brand.short_description = 'Brand'

    def product_sku(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    def product_kind(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    def color_variations(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    def text_variations(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_sku.short_description = 'Sku'

    def product_name(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_name.short_description = 'Name'

    def product_reference(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_reference.short_description = 'Barcode'

    def total_ordered(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    total_ordered.short_description = 'Total Ordered'

    def product_cost_price(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_cost_price.short_description = 'Cost Price'

    def product_quantity(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_quantity.short_description = 'Saved Stock'

    def in_transit_stock(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    in_transit_stock.short_description = 'In Transit Stock'

    def dc_stock(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    dc_stock.short_description = 'DC Stock'

    def product_snapshot_stock(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_snapshot_stock.short_description = 'DC Snapshot Stock'

    def product_snapshot_stock_date_time(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    product_snapshot_stock_date_time.short_description = 'DC Snapshot Stock Date Time'

    def difference_to_order(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    difference_to_order.admin_order_field = 'difference_to_order'
    difference_to_order.short_description = 'Difference To Order'

    def voucher_value(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''

    voucher_value.short_description = 'Voucher Value'

    def sent_to_approve(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return obj.get('sent_to_approve', 0)

    sent_to_approve.short_description = 'Sent to Approve'

    def voucher_type(self, obj):
        # this method isn't invoked, but is required for django to validate the
        # list_display fields. it returns a static value so as to not confuse
        # if the value ever needs to be changed (all values are computed in the
        # queryset above)
        return ''


@admin.register(RegularOrderSummary)
class RegularOrderSummaryAdmin(OrderSummaryAdmin):
    change_list_template = 'logistics/order_summary_change_list-regular-sent.html'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.GET.get('o') == '5':
            return qs.order_by('difference_to_order')
        if request.GET.get('o') == '-5':
            return qs.order_by('-difference_to_order')
        return qs


@admin.register(SupplierOrderSummary)
class SupplierOrderSummaryAdmin(OrderSummaryAdmin):
    change_list_template = 'logistics/order_summary_change_list-supplier-sent.html'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.GET.get('o') == '5':
            return qs.order_by('difference_to_order')
        if request.GET.get('o') == '-5':
            return qs.order_by('-difference_to_order')
        return qs


@admin.register(VoucherOrderSummary)
class VoucherOrderSummaryAdmin(OrderSummaryAdmin):
    change_list_template = 'logistics/order_summary_change_list-voucher.html'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.GET.get('o') == '5':
            return qs.order_by('difference_to_order')
        if request.GET.get('o') == '-5':
            return qs.order_by('-difference_to_order')
        return qs
