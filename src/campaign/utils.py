from io import BytesIO
import math
from typing import Optional

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db.models import (
    Case,
    CharField,
    DecimalField,
    ExpressionWrapper,
    F,
    FloatField,
    Max,
    OuterRef,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, Concat
from django.http import HttpResponse
import jwt
from openpyxl import Workbook
from openpyxl.chart import BarChart3D, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework.authentication import BaseAuthentication
from rest_framework.permissions import BasePermission

from campaign.models import (
    Campaign,
    CampaignEmployee,
    Employee,
    EmployeeGroup,
    EmployeeGroupCampaign,
    EmployeeGroupCampaignProduct,
    Order,
    OrganizationProduct,
    QuickOffer,
)
from inventory.models import Brand, Product, Tag, Variation


UserModel = get_user_model()


class EmployeeAuthentication(BaseAuthentication):
    def authenticate(self, request):
        try:
            auth = request.headers.get('X-Authorization')
            if isinstance(auth, str) and auth.startswith('Bearer '):
                token = auth.replace('Bearer ', '')
                if len(token) > 0:
                    decoded_token = jwt.decode(
                        jwt=token,
                        key=settings.JWT_SECRET_KEY,
                        algorithms=[settings.JWT_ALGORITHM],
                    )
                    if isinstance(decoded_token, dict):
                        if decoded_token.get('employee_id'):
                            employee = Employee.objects.get(
                                pk=decoded_token.get('employee_id')
                            )
                            return (employee, token)
                        elif decoded_token.get(
                            'impersonated_employee_id'
                        ) and decoded_token.get('admin_id'):
                            campaign_employee = CampaignEmployee.objects.get(
                                pk=decoded_token.get('impersonated_employee_id')
                            )
                            employee = campaign_employee.employee
                            impersonator_admin = UserModel.objects.get(
                                pk=decoded_token.get('admin_id')
                            )

                            _set_employee_impersonated_by(employee, impersonator_admin)

                            return (employee, token)
        except Exception:
            pass

        return None

    def authenticate_header(self, request):
        return 'Bearer'


class EmployeePermissions(BasePermission):
    def has_permission(self, request, view):
        return isinstance(request.user, Employee)


class AdminPreviewAuthentication(BaseAuthentication):
    def authenticate(self, request):
        try:
            auth = request.headers.get('X-Authorization')
            if isinstance(auth, str) and auth.startswith('Bearer '):
                token = auth.replace('Bearer ', '')
                if len(token) > 0:
                    decoded_token = jwt.decode(
                        jwt=token,
                        key=settings.JWT_SECRET_KEY,
                        algorithms=[settings.JWT_ALGORITHM],
                    )
                    if isinstance(decoded_token, dict):
                        if decoded_token.get('admin_preview') and decoded_token.get(
                            'admin_id'
                        ):
                            employee_group_campaign = EmployeeGroupCampaign.objects.get(
                                id=decoded_token.get('employee_group_campaign_id')
                            )

                            preview_employee = Employee(
                                employee_group=employee_group_campaign.employee_group,
                                first_name='Admin',
                                last_name='Admin',
                            )
                            _set_employee_admin_preview(preview_employee)

                            return (
                                preview_employee,
                                token,
                            )
        except Exception:
            pass

        return None

    def authenticate_header(self, request):
        return 'Bearer'


def get_campaign_product_price(
    campaign: Campaign, product: Product, employee_group: Optional[EmployeeGroup] = None
) -> int:
    if EmployeeGroup and campaign and product.product_kind == 'MONEY':
        c_product: EmployeeGroupCampaignProduct = (
            EmployeeGroupCampaignProduct.objects.filter(
                employee_group_campaign_id__campaign=campaign,
                employee_group_campaign_id__employee_group=employee_group,
                product_id=product,
            ).first()
        )

        if c_product:
            budget = c_product.employee_group_campaign_id.budget_per_employee or 0
            return budget

    organization_product = OrganizationProduct.objects.filter(
        organization=campaign.organization, product=product
    ).first()

    if organization_product and organization_product.price:
        return organization_product.price
    else:
        return product.sale_price if product.sale_price else 0


def get_quick_offer_product_price(quick_offer: QuickOffer, product: Product) -> int:
    organization_product = OrganizationProduct.objects.filter(
        organization=quick_offer.organization, product=product
    ).first()

    if organization_product and organization_product.price:
        return organization_product.price
    else:
        return product.sale_price if product.sale_price else 0


def _set_employee_admin_preview(employee: Employee) -> None:
    employee._admin_preview = True


def get_employee_admin_preview(employee: Employee) -> bool:
    return getattr(employee, '_admin_preview', False)


def _set_employee_impersonated_by(employee: Employee, admin_user: UserModel) -> None:
    employee._impersonated_by = admin_user.pk


def get_employee_impersonated_by(employee: Employee) -> bool:
    return getattr(employee, '_impersonated_by', None)


def format_with_none_replacement(format_string, **kwargs):
    cleaned_kwargs = {
        key: (
            f' ({value})'
            if key == 'delivery_additional_details' and value is not None
            else (
                f'{value}, '
                if (
                    (
                        key == 'delivery_street_number'
                        or key == 'delivery_apartment_number'
                    )
                    and value is not None
                )
                else (value if value is not None else '')
            )
        )
        for key, value in kwargs.items()
    }
    return format_string.format(**cleaned_kwargs)


def get_campaign_product_kinds(product_ids: list[int]):
    product_kinds = (
        Product.objects.filter(id__in=product_ids)
        .values_list('product_kind', flat=True)
        .distinct()
    )

    formatted_kinds = [
        {'id': product_kind.name, 'name': product_kind.value}
        for product_kind in [Product.ProductKindEnum[k] for k in product_kinds]
    ]

    return formatted_kinds


def get_campaign_brands(product_ids: list[int]):
    brands = Brand.objects.filter(brand_products__id__in=product_ids).distinct()

    return brands


def get_campaign_tags(product_ids: list[int]):
    tags = Tag.objects.filter(product__id__in=product_ids).distinct()

    return tags


def filter_campaign_products_with_calculated_price(
    campaign: Campaign, employee_group_campaign
):
    # Subquery to get product-related details
    employee_product_subquery = EmployeeGroupCampaignProduct.objects.filter(
        employee_group_campaign_id=employee_group_campaign,
        product_id=OuterRef('product_id'),
    ).values(
        'employee_group_campaign_id__budget_per_employee',
        'discount_mode',
        'organization_discount_rate',
    )[:1]

    # Annotate the queryset with the subquery results
    products = (
        EmployeeGroupCampaignProduct.objects.filter(
            employee_group_campaign_id=employee_group_campaign,
            product_id__active=True,
        )
        .select_related('product_id')
        .annotate(
            # Add the budget, discount_mode, and discount_rate as annotations
            budget_per_employee=Subquery(
                employee_product_subquery.values(
                    'employee_group_campaign_id__budget_per_employee'
                )[:1]
            ),
        )
        .annotate(
            calculated_price=Case(
                When(
                    product_id__product_kind='MONEY',
                    then=F('budget_per_employee'),
                ),
                default=Coalesce(
                    Subquery(
                        OrganizationProduct.objects.filter(
                            organization=campaign.organization,
                            product=OuterRef('product_id'),
                        ).values('price')[:1]
                    ),
                    F('product_id__sale_price'),
                    Value(0),
                    output_field=FloatField(),
                ),
                output_field=FloatField(),
            )
        )
    )

    return products


def get_campaign_max_product_price(
    product_ids: list[int], campaign: Campaign, employee_group_campaign
):
    max_price = (
        filter_campaign_products_with_calculated_price(
            campaign=campaign, employee_group_campaign=employee_group_campaign
        )
        .filter(product_id__in=product_ids)
        .aggregate(Max('calculated_price'))['calculated_price__max']
    )

    max_price = math.ceil(max_price) if max_price else 0

    return {'max_price': max_price}


def get_quick_offer_max_product_price(
    product_ids: list[int],
    quick_offer: QuickOffer,
    tax_percent: int,
):
    max_price = (
        Product.objects.filter(id__in=product_ids).annotate(
            calculated_price=Coalesce(
                Subquery(
                    OrganizationProduct.objects.filter(
                        organization=quick_offer.organization,
                        product=OuterRef('id'),
                    )
                    .annotate(
                        org_price=Case(
                            When(price__gt=0, then='price'),
                            default=None,
                            output_field=DecimalField(),
                        )
                    )
                    .values('org_price')[:1],  # subquery output
                ),
                'sale_price',
                output_field=DecimalField(),
            ),
            adjusted_price=ExpressionWrapper(
                F('calculated_price')
                / ((Value(100) + Value(tax_percent)) / Value(100.0)),
                output_field=DecimalField(),
            ),
        )
    ).aggregate(Max('adjusted_price'))['adjusted_price__max']

    max_price = math.ceil(max_price) if max_price else 0

    return {'max_price': max_price}


class QuickOfferAuthentication(BaseAuthentication):
    def authenticate(self, request):
        try:
            auth = str(request.headers.get('X-Authorization'))
            assert auth.startswith('Bearer ')
            token = auth.replace('Bearer ', '')
            assert len(token)
            decoded_token = jwt.decode(
                jwt=token,
                key=settings.JWT_SECRET_KEY,
                algorithms=[settings.JWT_ALGORITHM],
            )
            assert isinstance(decoded_token, dict)
            quick_offer_id = int(decoded_token.get('quick_offer_id'))
            quick_offer = QuickOffer.objects.get(id=quick_offer_id)
            setattr(request, 'quick_offer', quick_offer)
            return (None, token)
        except Exception:
            pass

        return None

    def authenticate_header(self, request):
        return 'Bearer'


class QuickOfferPermissions(BasePermission):
    def has_permission(self, request, view):
        return (
            isinstance(getattr(request, 'quick_offer', None), QuickOffer)
            and request.quick_offer.status == QuickOffer.StatusEnum.ACTIVE.name
        )


class EmployeeOrQuickOfferPermission(BasePermission):
    def has_permission(self, request, view):
        return QuickOfferPermissions().has_permission(
            request, view
        ) or EmployeePermissions().has_permission(request, view)


def transform_variations(variations, mode='response'):
    updated_variations = {}
    for variation_type, variation_value in variations.items():
        variation_instance = Variation.objects.filter(
            Q(site_name_he=variation_type) | Q(site_name_en=variation_type)
        ).first()
        if variation_instance:
            if mode == 'response':
                new_key = variation_instance.site_name_he or variation_type
            else:
                new_key = variation_instance.site_name_en or variation_type
            if (
                variation_instance.variation_kind
                == Variation.VariationKindEnum.TEXT.name
            ):
                text_variation_instance = variation_instance.text_variation.first()
                if text_variation_instance:
                    variation_value = (
                        text_variation_instance.text_he
                        if mode == 'response'
                        else text_variation_instance.text_en
                        or text_variation_instance.text_he
                    )
            updated_variations[new_key] = variation_value
        else:
            updated_variations[variation_type] = variation_value
    return updated_variations


def join_lists(list_1, list_2):
    list_1 = [item for item in list_1 if item]
    list_2 = [item for item in list_2 if item]
    list_1.extend(list_2)
    return list(set(list_1))


def get_campaign_employees(campaign: Campaign, export_type=None):
    campaign_employees = campaign.campaignemployee_set.select_related(
        'employee', 'employee__employee_group'
    ).annotate(
        full_name=Concat(
            'employee__first_name',
            Value(' '),
            'employee__last_name',
            output_field=CharField(),
        ),
        employee_group=F('employee__employee_group__name'),
    )
    employee_list = {}

    if export_type == 'employee_orders':
        order_data = {}
        pending_orders = Order.objects.filter(
            campaign_employee_id__campaign=campaign,
            status__in=[
                Order.OrderStatusEnum.PENDING.name,
                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
            ],
        ).order_by('-order_date_time')

        for orders_data in pending_orders:
            product_names = orders_data.ordered_product_names().split(', ')
            prev_names = (
                order_data.get(orders_data.campaign_employee_id.employee.id, {})
                .get('product_names', '')
                .split(', ')
            )
            prev_names = join_lists(prev_names, product_names)

            product_kinds = orders_data.ordered_product_kinds().split(', ')
            prev_kinds = (
                order_data.get(orders_data.campaign_employee_id.employee.id, {})
                .get('product_kinds', '')
                .split(', ')
            )
            prev_kinds = join_lists(prev_kinds, product_kinds)

            cost_added = orders_data.cost_added
            prev_cost_added = cost_added + order_data.get(
                orders_data.campaign_employee_id.employee.id, {}
            ).get('cost_added', 0)

            order_data.update(
                {
                    orders_data.campaign_employee_id.employee.id: {
                        'product_names': ', '.join(prev_names),
                        'product_kind': ', '.join(prev_kinds),
                        'order_date_time': orders_data.order_date_time,
                        'cost_added': prev_cost_added,
                        'extra_money': True if prev_cost_added else False,
                        'has_order': orders_data.order_date_time is not None,
                    }
                }
            )

        for employee in campaign_employees:
            if export_type == 'employee_orders':
                employee_list.update(
                    {
                        employee.employee.id: {
                            'employee_name': employee.full_name,
                            'employee_group': employee.employee_group,
                            'last_login': employee.last_login,
                            **order_data.get(
                                employee.employee.id,
                                {
                                    'product_names': '',
                                    'product_kind': '',
                                    'order_date_time': '',
                                    'added_cost': 0,
                                    'extra_money': False,
                                },
                            ),
                        }
                    }
                )

    if export_type == 'employee_budgets':
        employees = campaign_employees.annotate(
            group_budget=Subquery(
                EmployeeGroupCampaign.objects.filter(
                    employee_group=OuterRef('employee__employee_group'),
                    campaign=campaign,
                ).values('budget_per_employee')
            ),
            used_budget=Coalesce(
                Subquery(
                    Order.objects.filter(
                        campaign_employee_id__employee=OuterRef('employee'),
                        campaign_employee_id__campaign=campaign,
                        status__in=[
                            Order.OrderStatusEnum.PENDING.name,
                            Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                        ],
                    )
                    .values('campaign_employee_id')
                    .annotate(total_cost=Sum('cost_from_budget'))
                    .values('total_cost'),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
                Value(0, output_field=DecimalField(max_digits=10, decimal_places=2)),
            ),
            left_budget=ExpressionWrapper(
                F('total_budget') - F('used_budget'),
                output_field=DecimalField(max_digits=10, decimal_places=2),
            ),
        )
        for employee in employees:
            employee_list.update(
                {
                    employee.employee.id: {
                        'employee_name': employee.full_name,
                        'employee_group': employee.employee_group,
                        'total_budget': employee.total_budget,
                        'group_budget': employee.group_budget,
                        'used_budget': employee.used_budget,
                        'left_budget': employee.left_budget,
                    }
                }
            )

    return list(employee_list.values())


def get_xlsx_http_response(title: str, data: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    headers = list(data[0].keys())
    ws.append(headers)
    for d in data:
        ws.append([str(d.get(h)) for h in headers])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    return response


def get_xlsx_http_products_response(
    title: str, physical_products: list[dict], money_products: list[dict]
):
    wb = Workbook()
    ws_physical = wb.active
    ws_physical.title = 'Physical Products'
    if physical_products:  # Ensure there is data
        headers_physical = list(physical_products[0].keys())
        ws_physical.append(headers_physical)
        for d in physical_products:
            ws_physical.append([str(d.get(h)) for h in headers_physical])
    ws_money = wb.create_sheet(title='Money Products')
    if money_products:  # Ensure there is data
        headers_money = list(money_products[0].keys())
        ws_money.append(headers_money)
        for d in money_products:
            ws_money.append([str(d.get(h)) for h in headers_money])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    return response


def get_xlsx_http_campaign_response(
    title: str,
    main_data: list[dict],
    group_summaries_products: list[EmployeeGroupCampaign],
    organization: str,
    graph_categories: list[tuple],
    employee_group_selections: list[tuple],
    employees_choosing: list[tuple],
    choosing_by_day_time: list[tuple],
):
    wss = []
    wb = Workbook()
    ws_main_dimensions = {
        'A': 31,
        'B': 24,
        'C': 25,
        'D': 43,
        'E': 43,
        'F': 43,
        'G': 22,
        'H': 74,
        'I': 31,
    }
    ws_main = wb.active
    wss.append(ws_main)
    for column, width in ws_main_dimensions.items():
        ws_main.column_dimensions[column].width = width
    ws_main.sheet_view.rightToLeft = True
    ws_main.title = 'סיכום בחירות'
    headers_main = list(main_data[0].keys()) if main_data else []
    ws_main.append(headers_main)
    for d in main_data:
        ws_main.append([str(d.get(h)) for h in headers_main])
    for rows in ws_main.iter_rows(min_row=1, max_row=1, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(
                start_color='FFFF00', end_color='FFFF00', fill_type='solid'
            )

    ws_group_dimensions = {
        'A': 21,
        'B': 19,
        'C': 19,
        'D': 19,
        'E': 11,
    }

    group_summary_cost = {}
    for group_summary in group_summaries_products:
        ws_group = wb.create_sheet(title=f'{group_summary.employee_group.name} סיכום')
        for column, width in ws_group_dimensions.items():
            ws_group.column_dimensions[column].width = width
        wss.append(ws_group)
        ws_group.sheet_view.rightToLeft = True
        headers_group = [
            'פריט',
            'כמות מוצרים',
            'הנחה',
            'הנחה ל',
            'תקציב קבוצתי',
            'שווי שובר',
            'עלות כולל מעמ',
            'סה"כ עלות',
        ]
        ws_group.append([])
        ws_group.append([group_summary.employee_group.name])
        ws_group['A2'].font = Font(bold=True, underline='single')
        ws_group.append([])
        ws_group.append(headers_group)

        products_data = {}
        for product in group_summary.products:
            key = f'{product.product_kind}_{product.product_name}'
            if key in products_data.keys():
                continue
            products_data.update(
                {
                    key: {
                        'product_kind': product.product_kind,
                        'product_name': product.product_name,
                        'quantity': product.quantity,
                        'client_discount': product.client_discount
                        if product.client_discount
                        else 0,
                        'discount_to': product.discount_to,
                        'group_budget': product.group_budget,
                        'voucher_value': round(product.voucher_value, 1),
                        'cost_including_p': product.cost_including_p,
                        'total_cost': product.total_cost,
                    }
                }
            )

        money_product_total_cost = 0

        for product in products_data.values():
            if product.get('product_kind') == Product.ProductKindEnum.MONEY.name:
                money_product_total_cost = money_product_total_cost + product.get(
                    'total_cost'
                )
            else:
                continue

            ws_group.append(
                [
                    product.get('product_name'),
                    product.get('quantity'),
                    f'{product.get("client_discount")}%'
                    if product.get('client_discount')
                    else product.get('client_discount'),
                    product.get('discount_to'),
                    product.get('group_budget'),
                    product.get('voucher_value'),
                    product.get('cost_including_p'),
                    product.get('total_cost'),
                ]
            )

        group_summary_cost.update(
            {
                f'group_summary_{group_summary.pk}': {
                    'physical_product_total_cost': group_summary.physical_total_cost,
                    'money_product_total_cost': money_product_total_cost,
                }
            }
        )
        ws_group.append(['', '', '', '', ''])
        ws_group.append(
            [
                'סה"כ מתנות פיזיות',
                group_summary.physical_quantity,
                '',
                '',
                '',
                '',
                '',
                group_summary.physical_total_cost,
            ]
        )
        ws_group.append(
            [
                'סה"כ שוברים',
                group_summary.money_quantity,
                '',
                '',
                '',
                '',
                '',
                money_product_total_cost,
            ]
        )
        ws_group.append(['', '', '', '', ''])
        ws_group.append(
            [
                f'{group_summary.employee_group.name}סה"כ ',
                group_summary.all_quantity,
                '',
                '',
                '',
                '',
                '',
                money_product_total_cost + group_summary.physical_total_cost,
            ]
        )

        for rows in ws_group.iter_rows(min_row=4, max_row=4, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(
                    start_color='FFFF00', end_color='FFFF00', fill_type='solid'
                )

    ws_summary_dimensions = {'A': 26, 'B': 10, 'C': 22, 'D': 20}
    ws_summary = wb.create_sheet(title='סיכום כולל')
    for column, width in ws_summary_dimensions.items():
        ws_summary.column_dimensions[column].width = width
    wss.append(ws_summary)
    ws_summary.sheet_view.rightToLeft = True
    ws_summary.append([])
    ws_summary.append([])
    ws_summary.append([])

    headers_summary = ['קבוצה', 'פריט', 'כמות מוצרים', 'עלות כולל מעמ']
    ws_summary.append(headers_summary)

    total_money_products = 0
    total_physical_products = 0
    total_money_product_cost = 0
    total_physical_product_cost = 0
    for group_summary in group_summaries_products:
        total_physical_products = (
            total_physical_products + group_summary.physical_quantity
        )
        total_money_products = total_money_products + group_summary.money_quantity

        cost_dict = group_summary_cost.get(f'group_summary_{group_summary.pk}')
        physical_product_total_cost = cost_dict.get('physical_product_total_cost', 0)
        total_physical_product_cost = (
            total_physical_product_cost + physical_product_total_cost
        )
        money_product_total_cost = cost_dict.get('money_product_total_cost', 0)
        total_money_product_cost = total_money_product_cost + money_product_total_cost
        ws_summary.append(
            [
                group_summary.employee_group.name,
                'פיזי',
                group_summary.physical_quantity,
                physical_product_total_cost,
            ]
        )
        ws_summary.append(
            [
                group_summary.employee_group.name,
                'שוברים',
                group_summary.money_quantity,
                money_product_total_cost,
            ]
        )

    ws_summary.append(['', '', '', ''])
    ws_summary.append(
        [
            'סה"כ מתנות פיזיות',
            '',
            total_physical_products,
            total_physical_product_cost,
        ]
    )
    ws_summary.append(
        [
            'סה"כ שוברים',
            '',
            total_money_products,
            total_money_product_cost,
        ]
    )
    ws_summary.append(
        [
            'סה"כ',
            '',
            total_physical_products + total_money_products,
            total_physical_product_cost + total_money_product_cost,
        ]
    )

    for rows in ws_summary.iter_rows(min_row=4, max_row=4, min_col=1):
        for cell in rows:
            cell.fill = PatternFill(
                start_color='FFFF00', end_color='FFFF00', fill_type='solid'
            )

    ws_comments = wb.create_sheet(title='דרישת תשלום שוברים')
    ws_comments.column_dimensions['A'].width = 60
    ws_comments.column_dimensions['B'].width = 15
    ws_comments.sheet_view.rightToLeft = True

    ws_comments.append(['דרישת תשלום שוברים'])
    ws_comments['A1'].font = Font(bold=True, underline='single')
    ws_comments.append([f'{organization}: עבור'])
    ws_comments.append([])
    ws_comments.append([])
    ws_comments.append(['כמות שוברים', 'סהכ לתשלום'])
    ws_comments['A5'].font = Font(bold=True, underline='single')
    ws_comments['B5'].font = Font(bold=True, underline='single')
    ws_comments.append(
        [
            total_money_products,
            total_money_product_cost,
        ]
    )
    ws_comments.append([])
    ws_comments.append([])
    ws_comments.append([])

    ws_comments.append([': הערות'])
    ws_comments['A10'].font = Font(bold=True, underline='single')
    ws_comments['B10'].font = Font(bold=True, underline='single')
    ws_comments['A11'].value = 'נא לשלם עד שבוע ממועד קבלת דו"ח זה'
    ws_comments['A12'].value = 'נא לשלם לחשבון מספר 622993, סניף 723, בנק הפועלים 12'
    ws_comments.merge_cells(start_row=11, start_column=1, end_row=11, end_column=2)
    ws_comments.merge_cells(start_row=12, start_column=1, end_row=13, end_column=2)

    wss.append(ws_comments)
    ws_graphs_data = wb.create_sheet()
    ws_graphs = wb.create_sheet(title='גרפים')
    data = [
        ['Product Kind', 'Quantity'],
        [
            'Money Product',
            total_money_products,
        ],
        [
            'Physical Product',
            total_physical_products,
        ],
    ]
    for row in data:
        ws_graphs_data.append(row)

    pie = create_pie_chart(
        ws_graphs_data, 'Money Physical product selection', len(data)
    )

    ws_graphs.add_chart(pie, 'D1')

    # categories_graph
    ws_categories_graph_data = wb.create_sheet()
    data_categories = [*[('Category', 'Count')], *graph_categories]

    for row in data_categories:
        ws_categories_graph_data.append(row)

    chart1 = create_bar_chart(
        ws_categories_graph_data,
        'Category Selection Chart(Number)',
        len(data_categories),
    )

    ws_graphs.add_chart(chart1, 'N1')

    # Number of selections by group
    ws_selections_graph_data = wb.create_sheet()
    data_selections = [*[('Employee Group', 'Products')], *employee_group_selections]

    for row in data_selections:
        ws_selections_graph_data.append(row)

    pie_selections = create_pie_chart(
        ws_selections_graph_data, 'Number of selections by group', len(data_selections)
    )
    ws_graphs.add_chart(pie_selections, 'D21')

    # Choosing Rate
    ws_choosing_graph_data = wb.create_sheet()
    data_choosing = [*[('Type', 'Value')], *employees_choosing]

    for row in data_choosing:
        ws_choosing_graph_data.append(row)

    pie_choosing = create_pie_chart(
        ws_choosing_graph_data, 'Choosing Rate', len(data_choosing)
    )
    ws_graphs.add_chart(pie_choosing, 'N21')

    # choosing by day
    ws_choosing_by_day_graph_data = wb.create_sheet()
    data_choosing_by_day = [*[('Day', 'Count')], *choosing_by_day_time[0]]

    for row in data_choosing_by_day:
        ws_choosing_by_day_graph_data.append(row)

    chart_2 = create_bar_chart(
        ws_choosing_by_day_graph_data, 'Choosing by date', len(data_choosing_by_day)
    )

    ws_graphs.add_chart(chart_2, 'D40')

    # choosing by time
    ws_choosing_by_time_graph_data = wb.create_sheet()
    data_choosing_by_time = [*[('Time', 'Count')], *choosing_by_day_time[1]]

    for row in data_choosing_by_time:
        ws_choosing_by_time_graph_data.append(row)

    chart_3 = create_bar_chart(
        ws_choosing_by_time_graph_data, 'Choosing by time', len(data_choosing_by_time)
    )

    ws_graphs.add_chart(chart_3, 'N40')

    for ws in wss:
        for row in ws.iter_rows():
            for cell in row:
                if ws in [ws_main, ws_comments, ws_summary]:
                    alignment = Alignment(horizontal='center')
                    cell.alignment = alignment
                    continue
                alignment = Alignment(horizontal='right')
                cell.alignment = alignment

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
    return response


def price_deduct_tax(price: int) -> float:
    # calculate the price pre-tax
    return round(float(price) / ((100 + settings.TAX_PERCENT) / 100), 1)


def create_pie_chart(ws, title, max_row):
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=max_row)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    pie.title = title

    # Cut the first slice out of the pie
    _slice = DataPoint(idx=0, explosion=20)
    pie.series[0].data_points = [_slice]
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    ws.sheet_state = 'hidden'

    return pie


def create_bar_chart(ws, title, max_row):
    chart1 = BarChart3D()
    chart1.type = 'col'
    chart1.style = 10
    chart1.title = title

    chart1.x_axis.tickLblSkip = 1
    data_cat = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart1.add_data(data_cat, titles_from_data=True)
    chart1.legend = None
    chart1.set_categories(cats)

    chart1.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(
            anchor='ctr',
            anchorCtr='1',
            rot='-2700000',
            spcFirstLastPara='1',
            vertOverflow='ellipsis',
            wrap='square',
        ),
        p=[
            Paragraph(
                pPr=ParagraphProperties(defRPr=CharacterProperties()),
                endParaRPr=CharacterProperties(),
            )
        ],
    )
    ws.sheet_state = 'hidden'
    return chart1
