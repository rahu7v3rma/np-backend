from functools import update_wrapper
from io import BytesIO
import json
import logging
from typing import Any

from django.conf import settings
from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.db.models import (
    Case,
    CharField,
    Count,
    DateField,
    DecimalField,
    Exists,
    ExpressionWrapper,
    F,
    FilteredRelation,
    FloatField,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Cast, Coalesce, Concat, ExtractHour, Round
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils.datastructures import MultiValueDict
from django.utils.html import format_html
from django.utils.http import urlencode
from django.utils.safestring import mark_safe
from django.utils.translation import gettext_lazy as _
from django_admin_inline_paginator.admin import TabularInlinePaginated
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from campaign.tasks import send_campaign_employee_invitation
from inventory.models import Product
from inventory.utils import fill_message_template_email, fill_message_template_sms
from lib.admin import ImportableExportableAdmin, RecordImportError, custom_titled_filter
from lib.filters import MultiSelectFilter
from lib.models import StringAgg
from logistics.models import PurchaseOrder, PurchaseOrderProduct

from .admin_actions import (
    CampaignActionsMixin,
    OrderActionsMixin,
    QuickOfferActionsMixin,
)
from .admin_forms import (
    EmployeeForm,
    EmployeeGroupForm,
    ImportEmployeeGroupForm,
    ImportPricelistForm,
    OrderForm,
    OrderProductForm,
    OrderProductInlineForm,
    OrderProductInlineFormset,
)
from .admin_views import (
    CampaignCreationWizard,
    CampaignImpersonateView,
    CampaignInvitationView,
    QuickOfferCreationWizard,
    QuickOfferImpersonateView,
)
from .models import (
    Campaign,
    CampaignEmployee,
    DeliveryLocationEnum,
    Employee,
    EmployeeAuthEnum,
    EmployeeGroup,
    EmployeeGroupCampaign,
    EmployeeGroupCampaignProduct,
    Order,
    OrderProduct,
    Organization,
    OrganizationProduct,
    QuickOffer,
    QuickOfferTag,
)
from .serializers import (
    EmployeeReportSerializer,
    MoneyProductSerializerCampaignAdmin,
    PhysicalProductSerializerCampaignAdmin,
)
from .utils import (
    format_with_none_replacement,
    get_campaign_employees,
    get_xlsx_http_campaign_response,
    get_xlsx_http_products_response,
    get_xlsx_http_response,
)


logger = logging.getLogger(__name__)


class EmployeeGroupInline(admin.TabularInline):
    model = EmployeeGroup
    extra = 0


@admin.register(Organization)
class OrganizationAdmin(ImportableExportableAdmin):
    inlines = (EmployeeGroupInline,)

    fieldsets = [
        (
            None,
            {
                'fields': (
                    'name',
                    'manager_full_name',
                    'manager_phone_number',
                    'manager_email',
                    'logo_image',
                ),
            },
        ),
        ('ORGANIZATION PRODUCTS', {'fields': ('organization_products_link',)}),
    ]
    readonly_fields = ('organization_products_link',)

    actions = ['export_as_xlsx']
    change_list_template = 'admin/import_changelist.html'
    list_display = (
        'name',
        'manager_full_name',
        'number_of_employees',
    )
    add_form_template = 'admin/organization_form.html'
    change_form_template = 'campaign/change_organization_form.html'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.annotate(employees_count=Count('employeegroup__employee'))
        return qs

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj:
            self.organization_id = obj.id
        else:
            self.organization_id = 0
        return form

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        extra_context['organization_id'] = object_id
        extra_context['change'] = True

        if object_id and request.method == 'POST' and 'xlsx' in request.FILES:
            info = self.opts.app_label, self.opts.model_name
            viewname = 'admin:%s_%s_change' % info

            try:
                form = ImportPricelistForm(request.POST, request.FILES)
                if not form.is_valid():
                    self.message_user(
                        request,
                        'Imported pricelist file is invalid.',
                        level=messages.ERROR,
                    )
                    return redirect(reverse(viewname, args=(object_id,)))

                xlsx = form.clean().get('xlsx')
                workbook = load_workbook(xlsx)
                xlsx = workbook.active

                required_column_indices = {
                    'sku': None,
                    'organization price': None,
                }

                # read the first (title) row and fill required column indices
                for row in xlsx.iter_rows(max_row=1, max_col=10, values_only=True):
                    for i, cell in enumerate(row):
                        if cell in required_column_indices:
                            required_column_indices[cell] = i

                    break

                # if any required column was not found fail
                if any(v is None for v in required_column_indices.values()):
                    self.message_user(
                        request,
                        'An error occured while importing pricelist: '
                        'Imported sheet columns are invalid, '
                        'Download the sample by exporting pricelist.',
                        level=messages.ERROR,
                    )
                    return redirect(reverse(viewname, args=(object_id,)))

                org = Organization.objects.filter(pk=object_id).first()
                success = []
                errors = []

                for index, row in enumerate(
                    xlsx.iter_rows(min_row=2, max_col=10, values_only=True)
                ):
                    row_num = index + 2

                    empty_row = all(cell is None for cell in row)
                    if empty_row:
                        continue

                    # get product current values so that we don't depend on
                    # any value set in the excel file
                    sku = row[required_column_indices['sku']]
                    product = Product.objects.filter(sku=sku).first()

                    if not product:
                        errors.append(f'Could not find product with SKU "{sku}"')
                        continue

                    google_price = product.google_price
                    # total_cost = product.total_cost

                    org_price = row[required_column_indices['organization price']]
                    if org_price is not None:
                        if not isinstance(org_price, (float, int)):
                            errors.append(
                                f'Organization price is invalid for SKU ({sku}) at '
                                f'row number {row_num}.'
                            )
                            continue
                        else:
                            org_price = int(org_price)

                            # TODO: re-enable as a warning which requires
                            # confirmation to continue
                            # # validate organization price is not lower than total cost
                            # if org_price < total_cost:
                            #     errors.append(
                            #         'Organization price cannot be less than product '
                            #         f'total cost for SKU ({sku}) at row number '
                            #         f'{row_num}.'
                            #     )
                            #     continue

                    # validate organization price is not higher than google price
                    if (
                        google_price is not None
                        and org_price is not None
                        and org_price > google_price
                    ):
                        errors.append(
                            'Organization price cannot be higher than product google '
                            f'price for SKU ({sku}) at row number {row_num}.'
                        )
                        continue

                    org_product = OrganizationProduct.objects.filter(
                        organization=org,
                        product=product,
                    ).first()

                    # there are two cases - one is organization price is set to
                    # any value in which organization product must exist or be
                    # created, and one is organization price is None in which
                    # case if an organization product object exists we should
                    # remove it as the price was unset
                    if org_price is not None:
                        if org_product:
                            org_product.price = org_price
                            org_product.save()
                        else:
                            OrganizationProduct.objects.create(
                                organization=org,
                                product=product,
                                price=org_price,
                            )

                        success.append(sku)
                    elif org_price is None and org_product:
                        org_product.delete()
                        success.append(sku)

                if success:
                    message = 'Success updating organization price for products: '
                    message += ', '.join(success)
                    self.message_user(request, message, level=messages.SUCCESS)

                if errors:
                    errors.insert(
                        0, 'Errors updating organization price for following products:'
                    )
                    for message in errors:
                        self.message_user(request, message, level=messages.ERROR)

                return redirect(reverse(viewname, args=(object_id,)))
            except Exception as ex:
                logger.error(
                    f'An unknown error has occurred while importing prices: {ex}'
                )
                message = 'An unknown error has occurred while importing prices'
                self.message_user(request, message, level=messages.ERROR)
                return redirect(reverse(viewname, args=(object_id,)))

        return super(OrganizationAdmin, self).change_view(
            request, object_id, form_url, extra_context
        )

    def get_urls(self):
        urls = super().get_urls()
        app = self.opts.app_label
        model = self.opts.model_name
        custom_urls = [
            path(
                '<path:object_id>/export-pricelist/',
                self.export_pricelist,
                name=f'{app}_{model}_export_pricelist',
            ),
        ]
        return custom_urls + urls

    def export_pricelist(self, request, object_id):
        app = self.model._meta.app_label
        model = self.model._meta.model_name
        change_view = f'admin:{app}_{model}_change'

        try:
            org = Organization.objects.filter(pk=object_id).first()

            export_plain_columns = [
                'name_en',
                'name_he',
                'sku',
                'total_cost',
                'cost_price',
                'sale_price',
                'google_price',
                'product_kind',
                'supplier__name',
            ]

            export_calculated_columns = dict(
                organization_price=Cast(
                    'org_product__price', output_field=FloatField()
                ),
                bottom_margin=Round(
                    ExpressionWrapper(
                        F('cost_price') * Value(0.25), output_field=FloatField()
                    ),
                    2,
                ),
                top_margin=Round(
                    ExpressionWrapper(
                        F('cost_price') * Value(0.60), output_field=FloatField()
                    ),
                    2,
                ),
            )

            export_forumla_columns = {
                'profit': (
                    '=IF(ISBLANK($H{row_num}), "", ROUND(($H{row_num} '
                    '- $E{row_num}) / $E{row_num}, 2))'
                )
            }

            # used for forcing number formats on specific columns
            export_column_number_format = {
                'sku': '@',
                'profit': '0.00%',
            }

            products = Product.objects.annotate(
                org_product=FilteredRelation(
                    'organizationproduct',
                    condition=Q(organizationproduct__organization=org),
                ),
            ).values(*export_plain_columns, **export_calculated_columns)

            columns = (
                [c.replace('_', ' ') for c in export_plain_columns]
                + [c.replace('_', ' ') for c in export_calculated_columns.keys()]
                + list(export_forumla_columns.keys())
            )

            workbook = Workbook()
            xlsx = workbook.active
            xlsx.append(columns)

            for product_idx, product in enumerate(products):
                row = list(product.values())

                # add formula columns
                for formula in export_forumla_columns.values():
                    # this would be much easier if we could set a formula value
                    # to a range of cells, but it seems we can't so each
                    # cell formula should point to the correct row number
                    row.append(formula.format(row_num=product_idx + 2))

                xlsx.append(row)

            # force number formats where needed
            for k, v in export_column_number_format.items():
                # find column name in columns list and add 1 since worksheet
                # iteration values are 1-based
                column_index = columns.index(k) + 1

                for col in xlsx.iter_cols(
                    min_row=2, min_col=column_index, max_col=column_index
                ):
                    for cell in col:
                        cell.number_format = v

            red_fill = PatternFill(
                start_color='FF0000', end_color='FF0000', fill_type='solid'
            )
            green_fill = PatternFill(
                start_color='00FF00', end_color='00FF00', fill_type='solid'
            )
            no_fill = PatternFill(fill_type=None)

            # add conditional fill to organization price column values - no
            # fill for blank values, green fill if the value is between the
            # bottom and top margin and red fill otherwise ('notBetween'
            # operator did not work for some reason, so there's greaterThan and
            # lessThan instead)
            xlsx.conditional_formatting.add(
                f'H2:H{len(products) + 1}',
                CellIsRule(
                    operator='equal',
                    formula=['""'],
                    stopIfTrue=True,
                    fill=no_fill,
                ),
            )
            xlsx.conditional_formatting.add(
                f'H2:H{len(products) + 1}',
                CellIsRule(
                    operator='lessThan',
                    formula=['$I2'],
                    stopIfTrue=True,
                    fill=red_fill,
                ),
            )
            xlsx.conditional_formatting.add(
                f'H2:H{len(products) + 1}',
                CellIsRule(
                    operator='greaterThan',
                    formula=['$J2'],
                    stopIfTrue=True,
                    fill=red_fill,
                ),
            )
            xlsx.conditional_formatting.add(
                f'H2:H{len(products) + 1}',
                CellIsRule(
                    operator='between',
                    formula=['$I2', '$J2'],
                    stopIfTrue=True,
                    fill=green_fill,
                ),
            )

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = (
                f'attachment; filename={org.name}_pricelist.xlsx'
            )
            response.write(output.getvalue())
            return response
        except Exception as ex:
            logger.error(
                'An unknown error has occurred while exporting organization '
                f'prices: {ex}'
            )
            message = (
                'An unknown error has occurred while exporting organization prices'
            )
            self.message_user(request, message, level=messages.ERROR)
            return redirect(reverse(change_view, args=(object_id,)))

    def organization_products_link(self, obj):
        count = obj.products.count()

        organization_product_changelist_url = reverse(
            'admin:campaign_organizationproduct_changelist'
        )
        search_params = urlencode({'organization__id__exact': obj.id})

        return mark_safe(
            f'<a href="{organization_product_changelist_url}?{search_params}">'
            f'Total products: {count}</a>'
        )

    def number_of_employees(self, obj):
        return obj.employees_count

    number_of_employees.admin_order_field = 'employees_count'

    organization_products_link.short_description = 'Organization products'

    search_fields = ('name',)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['organization_names'] = json.dumps(
            list(Organization.objects.all().values_list('name', flat=True))
        )
        return super(OrganizationAdmin, self).changelist_view(request, extra_context)


@admin.register(Employee)
class EmployeeAdmin(ImportableExportableAdmin):
    list_display = (
        'first_name',
        'last_name',
        'employee_group',
        'organization',
        'email',
        'phone_number',
        'auth_id',
    )
    list_filter = ('employee_group__organization',)
    actions = ['export_as_xlsx']
    form = EmployeeForm
    change_form_template = 'admin/employee_form.html'
    change_list_template = 'campaign/employee_change_list.html'
    exclude = ('otp_secret',)
    search_fields = (
        'first_name',
        'last_name',
        'employee_group__name',
    )

    export_fields = (
        'id',
        'employee_group__name',
        'full_name_en',
        'full_name_he',
        'login_type',
        'auth_id',
        'email',
        'phone_number',
        'birthday_date',
        'default_language',
        'delivery_city',
        'delivery_street',
        'delivery_street_number',
        'delivery_apartment_number',
        'active',
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-data/organization/<int:organization_id>/',
                self.admin_site.admin_view(self.import_xlsx),
                name='import_data_organization_employees',
            ),
        ]
        return custom_urls + urls

    def replace_import_related_fields(self, record_field_values: dict):
        full_name_en = record_field_values.pop('full_name_en', None)
        if full_name_en:
            full_name = str(full_name_en).split(' ')
            first_name_en = str(full_name[0])
            last_name_en = str(' '.join(full_name[1:]))
            record_field_values.update(
                {'first_name_en': first_name_en, 'last_name_en': last_name_en}
            )

        full_name_he = record_field_values.pop('full_name_he', None)
        if full_name_he:
            full_name = str(full_name_he).split(' ')
            first_name_he = str(full_name[0])
            last_name_he = str(' '.join(full_name[1:]))
            record_field_values.update(
                {'first_name_he': first_name_he, 'last_name_he': last_name_he}
            )

        return record_field_values

    def import_parse_and_save_xlsx_data(
        self, extra_params: dict[str, Any], request_files: MultiValueDict
    ) -> tuple[int, int]:
        # load main xlsx file
        workbook = load_workbook(
            request_files.pop('xlsx_file')[0], data_only=True, read_only=True
        )
        worksheet = workbook.active

        columns = self.export_fields

        errors = {}

        records_created = 0
        records_updated = 0

        for row_idx, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                max_col=len(columns) + 1,
                values_only=True,
            )
        ):
            # break if we hit an empty row since this should be th end of the
            # sheet
            if not row or all(v is None for v in row):
                break

            record_field_values = {}

            try:
                shift = 0
                for col_idx, current_column in enumerate(columns):
                    if col_idx in self.import_excluded_fields_indexes:
                        shift += 1

                    if current_column == 'employee_group__name':
                        current_column = 'employee_group'
                    parsed_value = self.import_parse_field(
                        current_column,
                        row[col_idx + shift],
                        extra_params,
                        request_files,
                    )
                    record_field_values[current_column] = parsed_value

                record_field_values = self.replace_import_related_fields(
                    record_field_values=record_field_values
                )

                # the pk may or may not have been supplied
                record_pk = record_field_values.pop(self.model._meta.pk.name, None)

                if record_pk:
                    # update existing record
                    try:
                        record = self.model.objects.get(pk=record_pk)
                        for k, v in record_field_values.items():
                            setattr(record, k, v)
                        record.save()

                        records_updated += 1
                    # handle the case when the product is not found
                    except self.model.DoesNotExist:
                        record = self.model(**record_field_values)
                        record.full_clean()
                        record.save()

                        records_created += 1
                else:
                    # create record
                    record = self.model(**record_field_values)
                    record.full_clean()
                    record.save()

                    records_created += 1

            except ValidationError as ex:
                for mk in ex.message_dict.keys():
                    msg_details = ', '.join(
                        [m.removesuffix('.') for m in ex.message_dict[mk]]
                    )
                    msg = f'{mk} - {msg_details}'

                    if msg not in errors:
                        errors[msg] = []

                    # row_idx is 0-based, plus the first row is the field names
                    errors[msg].append(row_idx + 2)
            except (ValueError, Exception) as ex:
                # ValueErrors are raised by model-specific import logic (for
                # example product image matching). any other Exception should
                # also be caught here and handled the same way
                msg = str(ex)

                if msg not in errors:
                    errors[msg] = []

                # row_idx is 0-based, plus the first row is the field names
                errors[msg].append(row_idx + 2)

        if errors:
            raise RecordImportError(errors)
        else:
            return records_created, records_updated

    def import_parse_field(
        self,
        name: str,
        value: str,
        extra_params: dict[str, Any],
        extra_files: MultiValueDict,
    ):
        if name == 'employee_group':
            organization = Organization.objects.get(id=extra_params['organization_id'])
            return EmployeeGroup.objects.get(name=value, organization=organization)
        if name == 'full_name_en':
            super().import_parse_field(
                'first_name_en', value, extra_params, extra_files
            )
            return super().import_parse_field(
                'last_name_en', value, extra_params, extra_files
            )
        if name == 'full_name_he':
            super().import_parse_field(
                'first_name_he', value, extra_params, extra_files
            )
            return super().import_parse_field(
                'last_name_he', value, extra_params, extra_files
            )

        return super().import_parse_field(name, value, extra_params, extra_files)

    def save_model(self, request, obj, form, change):
        data = form.cleaned_data
        msgs = []
        if not data.get('organization'):
            msgs.append('Organization is required')
        if msgs:
            request.session['validation_error'] = True
            for msg in msgs:
                self.message_user(request, msg, level=messages.ERROR)
            return
        return super().save_model(request, obj, form, change)

    def response_change(self, request, obj):
        validation_error = request.session.get('validation_error')
        if validation_error:
            del request.session['validation_error']
            return redirect(request.path)
        return super().response_change(request, obj)

    def response_add(self, request, obj, post_url_continue=None):
        validation_error = request.session.get('validation_error')
        if validation_error:
            del request.session['validation_error']
            return redirect(request.path)
        return super().response_change(request, obj)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['employees'] = json.dumps(list())
        return super().changelist_view(request, extra_context)

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(
            request, queryset, search_term
        )
        if search_term:
            queryset = Employee.objects.filter(
                Q(first_name__icontains=search_term)
                | Q(last_name__icontains=search_term)
                | Q(employee_group__name__icontains=search_term)
                | Q(employee_group__campaigns__name__icontains=search_term)
                | Q(employee_group__organization__name__icontains=search_term)
            )
            use_distinct = True
        return queryset, use_distinct


class EmployeeInline(TabularInlinePaginated):
    model = Employee
    extra = 1
    exclude = ('otp_secret', 'first_name', 'last_name', 'total_budget')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        search_term = request.GET.get('q')
        if search_term:
            qs = qs.filter(
                Q(first_name__icontains=search_term)
                | Q(last_name__icontains=search_term)
                | Q(email__icontains=search_term)
                | Q(auth_id__icontains=search_term)
                | Q(phone_number__icontains=search_term)
            )
        return qs


@admin.register(EmployeeGroup)
class EmployeeGroupAdmin(admin.ModelAdmin):
    form = EmployeeGroupForm
    change_list_template = 'campaign/employee_group_change_list.html'
    change_form_template = 'admin/employee_group_change_form.html'
    list_display = ('name', 'campaign_names', 'organization', 'total_employees')
    inlines = [EmployeeInline]
    fields = (
        'name',
        'organization',
        'delivery_city',
        'delivery_street',
        'delivery_street_number',
        'delivery_apartment_number',
        'delivery_location',
        'auth_method',
    )

    search_fields = (
        'name',
        'organization__name',
    )

    def get_search_results(self, request, queryset, search_term):
        # Use the default search method first
        queryset, use_distinct = super().get_search_results(
            request, queryset, search_term
        )

        # Custom search for campaign names
        if search_term:
            campaign_name_search = Q(
                employeegroupcampaign__campaign__name__icontains=search_term
            )

            campaign_query_set = self.model.objects.filter(campaign_name_search)
            queryset = (queryset | campaign_query_set).distinct()  # Remove duplications

        return queryset, use_distinct

    def get_readonly_fields(self, request, obj=None):
        if obj:
            return self.readonly_fields + ('organization',)
        return self.readonly_fields

    actions = ['export_as_xlsx']
    export_fields = {
        'Id': 'id',
        'Name': 'name',
        'Organization': 'organization__name',
        'Delivery City': 'delivery_city',
        'Delivery Street': 'delivery_street',
        'Delivery Street Number': 'delivery_street_number',
        'Delivery Apartment Number': 'delivery_apartment_number',
        'Delivery Location': 'delivery_location',
        'Auth Method': 'auth_method',
    }

    def export_as_xlsx(self, request, queryset):
        data = list(queryset.values_list(*self.export_fields.values()))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=e=Employee-Groups.xlsx'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(list(self.export_fields.keys()))

        for _employee_group in data:
            worksheet.append(_employee_group)

        workbook.save(response)
        return response

    export_as_xlsx.short_description = 'Export selected Employee Group as XLSX'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-employee-group-xlsx/',
                self.admin_site.admin_view(self.import_xlsx_view),
                name='import_xlsx_view',
            ),
        ]
        return custom_urls + urls

    def import_xlsx_view(self, request):
        import_form = ImportEmployeeGroupForm(request.POST, request.FILES)
        if import_form.is_valid():
            try:
                employee_group_file = import_form.clean().get('employee_group_file')
                workbook = load_workbook(employee_group_file)
                sheet = workbook.active
                headers = None
                counter = 0
                for row in sheet.iter_rows(
                    max_col=len(self.export_fields), values_only=True
                ):
                    if not headers:
                        headers = row
                        continue

                    (
                        _id,
                        name,
                        organization__name,
                        delivery_city,
                        delivery_street,
                        delivery_street_number,
                        delivery_apartment_number,
                        delivery_location,
                        auth_method,
                    ) = row

                    organization = Organization.objects.filter(
                        name=organization__name
                    ).first()

                    employee_group = EmployeeGroup.objects.filter(id=_id).first()
                    if employee_group:
                        employee_group.name = name
                        employee_group.delivery_city = delivery_city
                        employee_group.delivery_street = delivery_street
                        employee_group.delivery_street_number = delivery_street_number
                        employee_group.delivery_apartment_number = (
                            delivery_apartment_number
                        )
                        employee_group.delivery_location = delivery_location
                        employee_group.auth_method = auth_method
                        employee_group.save()

                    else:
                        EmployeeGroup.objects.create(
                            name=name,
                            organization=organization,
                            delivery_city=delivery_city,
                            delivery_street=delivery_street,
                            delivery_street_number=delivery_street_number,
                            delivery_apartment_number=delivery_apartment_number,
                            delivery_location=delivery_location,
                            auth_method=auth_method,
                        )

                    counter = counter + 1

                self.message_user(
                    request,
                    f'successfully imported employee groups ({counter})',
                    level=messages.INFO,
                )
                return HttpResponse(status=200)
            except Exception as ex:
                print(ex)
                logger.error(
                    f'An unknown error has occurred while importing prices: {ex}'
                )
                message = (
                    'An unknown error has occurred while importing employee groups'
                )
                self.message_user(request, message, level=messages.ERROR)
                return HttpResponse(status=400)

    class Media:
        js = ('js/admin.js',)


@admin.register(Campaign)
class CampaignAdmin(admin.ModelAdmin, CampaignActionsMixin):
    actions = [
        'activate_campaign',
        'preview_campaign',
        'pending_approval_campaign',
        'finish_campaign',
        'export_orders_as_xlsx',
        'resend_invitation',
    ]

    list_display = (
        'name',
        'duplicate_link',
        'status',
        'total_employees',
        'ordered_number',
        'ordered_percentage',
        'organization_link',
        'employee_site_link',
    )

    search_fields = (
        'name_en',
        'name_he',
        'status',
        'organization__name_en',
        'organization__name_he',
    )

    list_filter = ('tags',)
    change_list_template = 'campaign/campaign_change_list.html'

    def employee_site_link(self, obj):
        href_str_list = []
        egc_obs = EmployeeGroupCampaign.objects.filter(campaign=obj)
        for egc in egc_obs:
            href_str_list.append(
                f'{egc.employee_group.name} - {egc.employee_site_link}'
            )

        link = '    --------    '.join(href_str_list)
        return format_html(
            '<a href="#" onclick="copyToClipboard(\'{0}\');'
            'return false;">Copy link</a>'
            '<input type="hidden" class="copy-link-input" value="{0}">',
            link,
        )

    def get_urls(self):
        app_label = self.opts.app_label
        model_name = self.opts.model_name
        return [
            path(
                '',
                update_wrapper(
                    self.admin_site.admin_view(self.changelist_view),
                    self.changelist_view,
                ),
                name=f'{app_label}_{model_name}_changelist',
            ),
            path(
                'add/',
                self.admin_site.admin_view(CampaignCreationWizard.as_view()),
                name=f'{app_label}_{model_name}_add',
            ),
            path(
                '<path:object_id>/change/',
                self.admin_site.admin_view(self.change_view),
                name=f'{app_label}_{model_name}_change',
            ),
            path(
                'edit/',
                self.admin_site.admin_view(CampaignCreationWizard.as_view()),
                name=f'{app_label}_{model_name}_edit',
            ),
            path(
                '<path:object_id>/status/',
                self.admin_site.admin_view(self.status_view),
                name=f'{app_label}_{model_name}_change',
            ),
            path(
                '<int:campaign_id>/preview/<int:employee_group_campaign_id>',
                CampaignImpersonateView.as_view(),
                name='campaign_preview_view',
            ),
            path(
                '<int:campaign_id>/impersonate/<int:campaign_employee_id>',
                CampaignImpersonateView.as_view(),
                name='campaign_impersonate_view',
            ),
            path(
                '<int:campaign_id>/invitation',
                CampaignInvitationView.as_view(),
                name='campaign_invitation_view',
            ),
        ]

    def get_campaign_products(self, campaign: Campaign):
        unique_product_ids = list(
            EmployeeGroupCampaignProduct.objects.filter(
                employee_group_campaign_id__campaign=campaign
            )
            .values_list('product_id', flat=True)
            .distinct()
        )

        money_product_list = Product.objects.filter(
            id__in=unique_product_ids, product_kind='MONEY'
        )

        physical_product_list = Product.objects.filter(
            id__in=unique_product_ids
        ).exclude(product_kind='MONEY')

        money_serializer = MoneyProductSerializerCampaignAdmin(
            money_product_list,
            many=True,
            context={'campaign': campaign},
        )

        physical_serializer = PhysicalProductSerializerCampaignAdmin(
            physical_product_list,
            many=True,
            context={'campaign': campaign},
        )

        return money_serializer.data, physical_serializer.data

    def get_campaign_employee_products(self, campaign: Campaign):
        campaign_employees = CampaignEmployee.objects.filter(
            campaign=campaign
        ).annotate(
            selection_date=Subquery(
                Order.objects.filter(
                    campaign_employee_id=OuterRef('pk'),
                    status__in=[
                        Order.OrderStatusEnum.PENDING.name,
                        Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                    ],
                )
                .order_by('-order_date_time')
                .values('order_date_time')[:1]
            ),
            has_order=Exists(
                Order.objects.filter(
                    campaign_employee_id=OuterRef('pk'),
                    status__in=[
                        Order.OrderStatusEnum.PENDING.name,
                        Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                    ],
                )
            ),
        )
        he_data = []
        for campaign_employee in campaign_employees:
            employee_serializer = EmployeeReportSerializer(
                campaign_employee.employee
            ).data
            if campaign_employee.has_order:
                products = OrderProduct.objects.filter(
                    order_id__campaign_employee_id=campaign_employee,
                    order_id__status__in=[
                        Order.OrderStatusEnum.PENDING.name,
                        Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                    ],
                    quantity__gt=0,
                ).annotate(name_he=F('product_id__product_id__name_he'))

                products_dict = {}

                for product in products:
                    name = product.name_he
                    quantity = product.quantity
                    products_dict.update({name: quantity + products_dict.get(name, 0)})

                for name, quantity in products_dict.items():
                    he_data.append(
                        {
                            'סוג בחירה': 'רגיל',
                            'תאריך בחירה': campaign_employee.selection_date.strftime(
                                '%Y-%m-%d %H:%M'
                            )
                            if campaign_employee.selection_date
                            else None,
                            'התחברות אחרונה': campaign_employee.last_login.strftime(
                                '%Y-%m-%d %H:%M'
                            )
                            if campaign_employee.last_login
                            else None,
                            'שם העובד': f'{employee_serializer.get("first_name_he")} '
                            + f'{employee_serializer.get("last_name_he")}',
                            'מייל העובד': employee_serializer.get('email'),
                            'טלפון עובד': employee_serializer.get('phone_number'),
                            'קבוצת עובד': employee_serializer.get(
                                'employee_group', {}
                            ).get('name'),
                            'שם המוצר': name if name else None,
                            ' מספר מוצרים שנבחרו': quantity if quantity else None,
                        }
                    )
            else:
                # If employee has no orders, add them with empty product fields
                he_data.append(
                    {
                        'סוג בחירה': 'רגיל',
                        'תאריך בחירה': None,  # Since the employee hasn't order
                        'התחברות אחרונה': campaign_employee.last_login.strftime(
                            '%Y-%m-%d %H:%M'
                        )
                        if campaign_employee.last_login
                        else None,
                        'שם העובד': f'{employee_serializer.get("first_name_he")} {employee_serializer.get("last_name_he")}',  # noqa: E501
                        'מייל העובד': employee_serializer.get('email'),
                        'טלפון עובד': employee_serializer.get('phone_number'),
                        'קבוצת עובד': employee_serializer.get('employee_group', {}).get(
                            'name'
                        ),
                        'שם המוצר': None,
                        ' מספר מוצרים שנבחרו': None,
                    }
                )
        return he_data

    def get_group_summaries_products(self, campaign: Campaign):
        order_products = self.get_order_products(campaign=campaign).filter(
            quantity__gt=0,
        )
        data = (
            EmployeeGroupCampaign.objects.filter(campaign=campaign)
            .prefetch_related(
                Prefetch(
                    'employeegroupcampaignproduct_set',
                    queryset=EmployeeGroupCampaignProduct.objects.filter(
                        orderproduct__id__in=order_products.values_list(
                            'id', flat=True
                        ),
                    )
                    .annotate(
                        product_name=F('product_id__name_he'),
                        product_kind=F('product_id__product_kind'),
                        discount_to=F('discount_mode'),
                        group_budget=F(
                            'employee_group_campaign_id__budget_per_employee'
                        ),
                        organization_discount=Coalesce(
                            F('organization_discount_rate'),
                            Value(0),
                            output_field=FloatField(),
                        ),
                        voucher_value=Coalesce(
                            Case(
                                When(
                                    discount_mode=EmployeeGroupCampaign.DefaultDiscountTypeEnum.EMPLOYEE.name,
                                    then=Case(
                                        When(
                                            organization_discount=Value(0),
                                            then=F('group_budget'),
                                        ),
                                        default=F('group_budget')
                                        / (
                                            Value(1)
                                            - (F('organization_discount') / Value(100))
                                        ),
                                        output_field=FloatField(),
                                    ),
                                ),
                                default=F('group_budget'),
                                output_field=FloatField(),
                            ),
                            Value(0),
                            output_field=FloatField(),
                        ),
                        cost_including_p=Case(
                            When(
                                discount_mode=EmployeeGroupCampaign.DefaultDiscountTypeEnum.ORGANIZATION.name,
                                then=F('group_budget')
                                * (
                                    Value(1)
                                    - (
                                        Coalesce(
                                            F('organization_discount_rate'), Value(0)
                                        )
                                        / Value(100)
                                    )
                                ),
                            ),
                            default=F('group_budget'),
                            output_field=FloatField(),
                        ),
                        client_discount=F('organization_discount_rate'),
                        quantity=Coalesce(
                            Subquery(
                                OrderProduct.objects.filter(
                                    product_id=OuterRef('pk'),
                                    order_id__status__in=[
                                        Order.OrderStatusEnum.PENDING.name,
                                        Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                                    ],
                                )
                                .values('product_id')
                                .annotate(
                                    has_physical=Subquery(
                                        OrderProduct.objects.filter(
                                            order_id__campaign_employee_id=OuterRef(
                                                'order_id__campaign_employee_id'
                                            ),
                                            product_id__product_id__product_kind__in=[
                                                'PHYSICAL',
                                                'BUNDLE',
                                                'VARIATION',
                                            ],
                                            order_id__status__in=[
                                                Order.OrderStatusEnum.PENDING.name,
                                                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                                            ],
                                        )
                                        .values('order_id')
                                        .annotate(count=Count('id'))
                                        .values('count')[:1]
                                    )
                                )
                                .filter(
                                    Q(has_physical__isnull=True)
                                    | Q(
                                        has_physical__gt=0,
                                        product_id__product_id__product_kind__in=[
                                            'PHYSICAL',
                                            'BUNDLE',
                                            'VARIATION',
                                        ],
                                    )
                                )
                                .annotate(
                                    total_unique_employees=Count(
                                        'order_id__campaign_employee_id__employee',
                                        distinct=True,
                                    ),
                                )
                                .values('total_unique_employees')[:1],
                            ),
                            Value(0),
                        ),
                        total_cost=Case(
                            When(
                                product_id__product_kind='MONEY',
                                then=F('quantity') * F('cost_including_p'),
                            ),
                            default=Coalesce(
                                Subquery(
                                    OrderProduct.objects.filter(
                                        product_id=OuterRef('pk'),
                                        order_id__status__in=[
                                            Order.OrderStatusEnum.PENDING.name,
                                            Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                                        ],
                                    )
                                    .values('order_id__campaign_employee_id__employee')
                                    .distinct()
                                    .annotate(
                                        employee_budget=F(
                                            'order_id__campaign_employee_id__total_budget'
                                        )
                                    )
                                    .values('product_id')
                                    .annotate(
                                        employee_total_budget=Sum('employee_budget')
                                    )
                                    .values('employee_total_budget')[:1]
                                ),
                                Value(0),
                            ),
                            output_field=DecimalField(max_digits=10, decimal_places=2),
                        ),
                    )
                    .filter(quantity__gt=0),
                    to_attr='products',
                )
            )
            .annotate(
                physical_quantity=Coalesce(
                    Subquery(
                        OrderProduct.objects.filter(
                            product_id__employee_group_campaign_id=OuterRef('pk'),
                            order_id__status__in=[
                                Order.OrderStatusEnum.PENDING.name,
                                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                            ],
                            product_id__product_id__product_kind__in=[
                                'PHYSICAL',
                                'BUNDLE',
                                'VARIATION',
                            ],
                        )
                        .annotate(
                            product_category=Case(
                                When(
                                    product_id__product_id__product_kind='MONEY',
                                    then=Value('MONEY'),
                                ),
                                default=Value('OTHER'),
                                output_field=CharField(),
                            )
                        )
                        .values('product_category')
                        .annotate(
                            total_unique_employees=Count(
                                'order_id__campaign_employee_id__employee',
                                distinct=True,
                            )
                        )
                        .values('total_unique_employees')[:1],
                    ),
                    Value(0),
                ),
                money_quantity=Coalesce(
                    Subquery(
                        OrderProduct.objects.filter(
                            product_id__employee_group_campaign_id=OuterRef('pk'),
                            order_id__status__in=[
                                Order.OrderStatusEnum.PENDING.name,
                                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                            ],
                            product_id__product_id__product_kind__in=['MONEY'],
                        )
                        .annotate(
                            has_physical=Subquery(
                                OrderProduct.objects.filter(
                                    order_id__campaign_employee_id=OuterRef(
                                        'order_id__campaign_employee_id'
                                    ),
                                    product_id__product_id__product_kind__in=[
                                        'PHYSICAL',
                                        'BUNDLE',
                                        'VARIATION',
                                    ],
                                    order_id__status__in=[
                                        Order.OrderStatusEnum.PENDING.name,
                                        Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                                    ],
                                )
                                .values('order_id')
                                .annotate(count=Count('id'))
                                .values('count')[:1]
                            )
                        )
                        .filter(
                            Q(has_physical__isnull=True)
                            | Q(
                                has_physical__gt=0,
                                product_id__product_id__product_kind__in=[
                                    'PHYSICAL',
                                    'BUNDLE',
                                    'VARIATION',
                                ],
                            )
                        )
                        .annotate(
                            product_category=Case(
                                When(
                                    product_id__product_id__product_kind='MONEY',
                                    then=Value('MONEY'),
                                ),
                                default=Value('OTHER'),
                                output_field=CharField(),
                            )
                        )
                        .values('product_category')
                        .annotate(
                            total_unique_employees=Count(
                                'order_id__campaign_employee_id__employee',
                                distinct=True,
                            )
                        )
                        .values('total_unique_employees')[:1],
                    ),
                    Value(0),
                ),
                all_quantity=Coalesce(
                    Subquery(
                        OrderProduct.objects.filter(
                            product_id__employee_group_campaign_id=OuterRef('pk'),
                            order_id__status__in=[
                                Order.OrderStatusEnum.PENDING.name,
                                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                            ],
                        )
                        .annotate(product_category=Value('ALL'))
                        .values('product_category')
                        .annotate(
                            total_unique_employees=Count(
                                'order_id__campaign_employee_id__employee',
                                distinct=True,
                            )
                        )
                        .values('total_unique_employees')[:1],
                    ),
                    Value(0),
                ),
                physical_total_cost=Coalesce(
                    Subquery(
                        # different approach with CampaignEmployee ,
                        # because distinct is not working with sqlite
                        CampaignEmployee.objects.filter(
                            campaign=OuterRef('campaign'),
                            employee__employee_group=OuterRef('employee_group'),
                            # Only include employees who have physical products
                            id__in=Subquery(
                                OrderProduct.objects.filter(
                                    product_id__employee_group_campaign_id__employee_group=OuterRef(
                                        'employee__employee_group'
                                    ),
                                    order_id__status__in=[
                                        Order.OrderStatusEnum.PENDING.name,
                                        Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                                    ],
                                    product_id__product_id__product_kind__in=[
                                        'PHYSICAL',
                                        'BUNDLE',
                                        'VARIATION',
                                    ],
                                )
                                .values('order_id__campaign_employee_id')
                                .distinct()
                            ),
                        )
                        .values('employee__employee_group')
                        # This distinct is not working with sqlite
                        # (Thats why I changed OrderProduct to CampaignEmployee)
                        .distinct()
                        .annotate(total_budget=Sum('total_budget'))
                        .values('total_budget')[:1]
                    ),
                    Value(0),
                ),
            )
        )

        return data

    def get_order_products(self, campaign):
        return OrderProduct.objects.filter(
            product_id__employee_group_campaign_id__campaign=campaign,
            order_id__status__in=[
                Order.OrderStatusEnum.PENDING.name,
                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
            ],
        )

    def get_graph_categories_data(self, campaign):
        order_products = self.get_order_products(campaign=campaign)
        return list(
            EmployeeGroupCampaignProduct.objects.filter(
                employee_group_campaign_id__campaign=campaign,
                product_id__in=order_products.values('product_id__product_id'),
                orderproduct__quantity__gt=0,
            )
            .exclude(product_id__categories__name=None)
            .values('product_id__categories__name')
            .annotate(Count('product_id__categories__name'))
            .values_list(
                'product_id__categories__name', 'product_id__categories__name__count'
            )
        )

    def get_graph_employee_group_selections(self, campaign: Campaign):
        order_products = self.get_order_products(campaign=campaign)
        return list(
            EmployeeGroupCampaignProduct.objects.filter(
                employee_group_campaign_id__campaign=campaign,
                product_id__in=order_products.values('product_id__product_id'),
                orderproduct__quantity__gt=0,
            )
            .values('employee_group_campaign_id__employee_group__name')
            .annotate(products=Count('product_id'))
            .values_list('employee_group_campaign_id__employee_group__name', 'products')
        )

    def get_graph_employee_choosing(self, campaign):
        order_products = self.get_order_products(campaign=campaign)
        choosing_employees = (
            order_products.values('order_id__campaign_employee_id__employee__id')
            .distinct()
            .count()
        )
        total_employees = CampaignEmployee.objects.filter(campaign=campaign).count()
        return [
            ('normal', choosing_employees),
            ('default', total_employees - choosing_employees),
        ]

    def get_choosing_by_day_time(self, campaign):
        order_products = self.get_order_products(campaign=campaign)
        choosing_by_day = list(
            order_products.annotate(
                day=Cast(Cast('order_id__order_date_time', DateField()), CharField())
            )
            .values('day')
            .annotate(selections=Count('product_id'))
            .order_by('-day')
            .values_list('day', 'selections')
        )

        choosing_by_time = list(
            order_products.values(f_time=ExtractHour('order_id__order_date_time'))
            .annotate(selections=Count('product_id'))
            .order_by('-f_time')
            .annotate(
                time=Concat(F('f_time') - 2, Value('h'), output_field=CharField())
            )
            .values_list('time', 'selections')
        )

        return choosing_by_day, choosing_by_time

    def status_view(self, request, object_id):
        campaign = Campaign.objects.get(pk=object_id)
        export_type = request.GET.get('export_type')

        if export_type in ['employee_orders', 'employee_budgets']:
            # Get data with appropriate export type
            campaign_employees = get_campaign_employees(campaign, export_type)

            # Set title based on export type
            if export_type == 'employee_orders':
                title = 'Campaign employee orders'
            else:
                title = 'Campaign employee budgets'

            xlsx_http_response = get_xlsx_http_response(
                title,
                campaign_employees,
            )
            return xlsx_http_response
        elif export_type == 'product':
            money_products, physical_products = self.get_campaign_products(campaign)
            xlsx_http_response = get_xlsx_http_products_response(
                title='Campaign products',
                physical_products=physical_products,
                money_products=money_products,
            )
            return xlsx_http_response

        elif export_type == 'campaign':
            group_summaries_products = self.get_group_summaries_products(campaign)
            main_data = self.get_campaign_employee_products(campaign)
            graph_categories = self.get_graph_categories_data(campaign)
            employee_group_selections = self.get_graph_employee_group_selections(
                campaign
            )
            employees_choosing = self.get_graph_employee_choosing(campaign)
            choosing_by_day_time = self.get_choosing_by_day_time(campaign)
            xlsx_http_response = get_xlsx_http_campaign_response(
                title='campaign selection',
                main_data=main_data,
                group_summaries_products=list(group_summaries_products),
                organization=campaign.organization.name,
                graph_categories=graph_categories,
                employee_group_selections=employee_group_selections,
                employees_choosing=employees_choosing,
                choosing_by_day_time=choosing_by_day_time,
            )
            return xlsx_http_response

        elif export_type == 'product':
            money_products, physical_products = self.get_campaign_products(campaign)
            xlsx_http_response = get_xlsx_http_products_response(
                title='Campaign products',
                physical_products=physical_products,
                money_products=money_products,
            )
            return xlsx_http_response

        send_invitation_type = request.GET.get('sendInvitationType')
        raw_ids = request.GET.get('campaignEmployees', '').split(',')
        campaign_employees = []
        for item in raw_ids:
            try:
                campaign_employees.append(int(item))
            except ValueError:
                pass

        campaign_employees = [
            item for item in campaign_employees if isinstance(item, int)
        ]
        send_invitation_campaign_employees = list(map(int, campaign_employees))

        campaign_employees = CampaignEmployee.objects.filter(
            campaign_id=object_id
        ).first()
        if not campaign_employees:
            context = {
                **self.admin_site.each_context(request),
                'opts': self.opts,
                'object_id': object_id,
                'title': 'Campaign Status',
                'subtitle': campaign.name,
                'status': None,
                'employee_groups': (
                    campaign.employeegroupcampaign_set.select_related(
                        'employee_group'
                    ).values('id', employee_group_name=F('employee_group__name'))
                ),
                'campaign_active': campaign.status
                in [
                    Campaign.CampaignStatusEnum.ACTIVE.name,
                    Campaign.CampaignStatusEnum.PREVIEW.name,
                ],
                'campaign_code': campaign.code,
            }
            return TemplateResponse(request, 'campaign/status_form.html', context)

        raw_status = (
            CampaignEmployee.objects.filter(campaign_id=object_id)
            .select_related('employee', 'employee__employee_group', 'order')
            .annotate(
                group_budget=Subquery(
                    EmployeeGroupCampaign.objects.filter(
                        employee_group=OuterRef('employee__employee_group'),
                        campaign=campaign,
                    ).values('budget_per_employee')
                ),
                total_budget_decimal=Cast(
                    'total_budget',
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
                employee_used_budget=Coalesce(
                    Subquery(
                        Order.objects.filter(
                            campaign_employee_id=OuterRef('pk'),
                            status__in=[
                                Order.OrderStatusEnum.PENDING.name,
                                Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                            ],
                        )
                        .values('campaign_employee_id')
                        .annotate(total_cost=Sum('cost_from_budget'))
                        .values('total_cost'),
                        output_field=DecimalField(max_digits=10, decimal_places=2),
                    ),
                    Value(
                        0, output_field=DecimalField(max_digits=10, decimal_places=2)
                    ),
                ),
                employee_left_budget=ExpressionWrapper(
                    F('total_budget_decimal') - F('employee_used_budget'),
                    output_field=DecimalField(max_digits=10, decimal_places=2),
                ),
                pending_order=FilteredRelation(
                    'order',
                    condition=Q(
                        order__status__in=[
                            Order.OrderStatusEnum.PENDING.name,
                            Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                            Order.OrderStatusEnum.COMPLETE.name,
                        ]
                    ),
                ),
                link=Case(
                    When(
                        employee__login_type=EmployeeAuthEnum.SMS.name,
                        then=Concat(
                            Value(settings.EMPLOYEE_SITE_BASE_URL),
                            Value('/'),
                            Value(campaign.code),
                            Value('/'),
                            Value('p'),
                        ),
                    ),
                    When(
                        employee__login_type=EmployeeAuthEnum.AUTH_ID.name,
                        then=Concat(
                            Value(settings.EMPLOYEE_SITE_BASE_URL),
                            Value('/'),
                            Value(campaign.code),
                            Value('/'),
                            Value('a'),
                        ),
                    ),
                    default=Concat(
                        Value(settings.EMPLOYEE_SITE_BASE_URL),
                        Value('/'),
                        Value(campaign.code),
                        Value('/'),
                        Value('e'),
                    ),
                ),
            )
            .values(
                link=F('link'),
                emp_id=F('employee_id'),
                employee_group_id=F('employee__employee_group__id'),
                campaign_employee_id=F('pk'),
                campaign_employee_last_login=F('last_login'),
                group_budget=F('group_budget'),
                employee_total_budget=F('total_budget_decimal'),
                employee_used_budget=F('employee_used_budget'),
                left_budget=F('employee_left_budget'),
                employee_group_name=F('employee__employee_group__name'),
                employee_first_name=F('employee__first_name'),
                employee_last_name=F('employee__last_name'),
                order_id=StringAgg(
                    Cast('pending_order__pk', output_field=CharField()),
                    Value(', '),
                    distinct=True,
                ),
                ordered_products=StringAgg(
                    F('pending_order__orderproduct__product_id__product_id__name'),
                    Value(', '),
                ),
                order_date_time=StringAgg(
                    Cast('pending_order__order_date_time', output_field=CharField()),
                    Value(', '),
                    distinct=True,
                ),
            )
            .order_by('order_id', 'emp_id')
        )

        if send_invitation_type and send_invitation_campaign_employees:
            send_invitation_employees = list(
                CampaignEmployee.objects.filter(
                    id__in=send_invitation_campaign_employees
                ).values_list('employee__id', flat=True)
            )
            send_campaign_employee_invitation.apply_async(
                (object_id, send_invitation_employees, send_invitation_type)
            )
            self.message_user(
                request, 'Invitation sending process has started', messages.SUCCESS
            )

        context = {
            **self.admin_site.each_context(request),
            'opts': self.opts,
            'object_id': object_id,
            'title': 'Campaign Status',
            'subtitle': campaign.name,
            'status': list(raw_status),
            'employee_groups': campaign.employeegroupcampaign_set.select_related(
                'employee_group'
            ).values('id', employee_group_name=F('employee_group__name')),
            'campaign_active': campaign.status
            in [
                Campaign.CampaignStatusEnum.ACTIVE.name,
                Campaign.CampaignStatusEnum.PREVIEW.name,
            ],
            'campaign_code': campaign.code,
            'sms_sender_name': campaign.sms_sender_name,
            'sms_welcome_text': fill_message_template_sms(
                employee=campaign_employees.employee, campaign=campaign
            ),
            'email_welcome_text': fill_message_template_email(
                employee=campaign_employees.employee, campaign=campaign
            ),
        }
        return TemplateResponse(request, 'campaign/status_form.html', context)

    def duplicate_link(self, obj):
        return mark_safe(
            '<a href="%s?campaign=%s&duplicate=true">duplicate</a>'
            % (reverse('admin:campaign_campaign_add'), obj.id)
        )

    class Media:
        js = ('js/admin.js',)


class OrderProductInline(admin.TabularInline):
    model = OrderProduct
    formset = OrderProductInlineFormset
    extra = 0
    form = OrderProductInlineForm
    fields = [
        'purchase_order_product',
        'product_id',
        'sku_display',
        'supplier_display',
        'organization_price',
        'cost_price',
        'voucher_val',
        'quantity',
        'variations',
    ]
    readonly_fields = [
        'sku_display',
        'supplier_display',
        'organization_price',
        'cost_price',
        'voucher_val',
    ]

    class Media:
        js = ('js/order_inline.js',)
        css = {'all': ('css/order_inline.css',)}

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        if obj:
            # Get all possible products for this campaign/employee group
            products = EmployeeGroupCampaignProduct.objects.filter(
                employee_group_campaign_id__campaign=obj.campaign_employee_id.campaign,
                employee_group_campaign_id__employee_group=obj.campaign_employee_id.employee.employee_group,
                active=True,
            ).select_related('product_id__supplier')

            # Create a dictionary of product data
            product_data = {
                str(p.id): {
                    'sku': p.product_id.sku,
                    'supplier': p.product_id.supplier.name
                    if p.product_id.supplier
                    else '',
                }
                for p in products
            }
            # Add the product data to the form's widget
            formset.form.base_fields['product_id'].widget.attrs['data-products'] = (
                json.dumps(product_data)
            )
            formset.form.base_fields['product_id'].widget.url_params = {
                'campaign_id': obj.campaign_employee_id.campaign.id,
                'employee_group_id': obj.campaign_employee_id.employee.employee_group.id,  # noqa: E501
            }
        return formset

    def sku_display(self, obj):
        if obj and obj.product_id and obj.product_id.product_id:
            return obj.product_id.product_id.sku
        return ''

    sku_display.short_description = 'SKU'

    def supplier_display(self, obj):
        if (
            obj
            and obj.product_id
            and obj.product_id.product_id
            and obj.product_id.product_id.supplier
        ):
            return obj.product_id.product_id.supplier.name
        return ''

    supplier_display.short_description = 'Supplier'

    def organization_price(self, obj):
        if (
            obj
            and obj.product_id
            and obj.product_id.product_id
            and obj.product_id.product_id.sale_price
        ):
            return obj.product_id.product_id.sale_price
        return ''

    organization_price.short_description = 'Organization Price'

    def cost_price(self, obj):
        if (
            obj
            and obj.product_id
            and obj.product_id.product_id
            and obj.product_id.product_id.cost_price
        ):
            return obj.product_id.product_id.cost_price
        return ''

    cost_price.short_description = 'Cost Price'

    def voucher_val(self, obj):
        if (
            obj
            and obj.purchase_order_product
            and obj.purchase_order_product.voucher_value
        ):
            return obj.purchase_order_product.voucher_value
        return ''

    voucher_val.short_description = 'Voucher Value'


class CampaignFilter(admin.SimpleListFilter):
    title = 'Campaign'
    parameter_name = 'campaign'

    def lookups(self, request, model_admin):
        return [(nm, nm) for nm in Campaign.objects.values_list('name', flat=True)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(
                campaign_employee_id__campaign__name__in=self.value().split(',')
            )
        return queryset


class OrganizationFilter(admin.SimpleListFilter):
    title = 'Organization'
    parameter_name = 'organization'

    def lookups(self, request, model_admin):
        return [(nm, nm) for nm in Organization.objects.values_list('name', flat=True)]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(
                campaign_employee_id__campaign__organization__name__in=self.value().split(
                    ','
                )
            )
        return queryset


class OrderProductFilter(admin.SimpleListFilter):
    title = 'Product ID'
    parameter_name = 'product_id'

    def lookups(self, request, model_admin):
        items = [
            (product_id, product_id)
            for product_id in OrderProduct.objects.values_list(
                'product_id__product_id', flat=True
            )
        ]
        filter_data = []
        for item in items:
            if item not in filter_data:
                filter_data.append(item)
        return filter_data

    def queryset(self, request, queryset):
        purchase_order_id = request.GET.get('purchase_order')
        if self.value() and not purchase_order_id:
            return queryset.filter(
                pk__in=OrderProduct.objects.filter(
                    product_id__product_id__in=str(self.value()).split(','),
                    purchase_order_product__isnull=True,
                    order_id__status__in=[
                        Order.OrderStatusEnum.PENDING.name,
                        Order.OrderStatusEnum.SENT_TO_LOGISTIC_CENTER.name,
                    ],
                ).values('order_id')
            )
        return queryset


class PurchaseOrderFilter(admin.SimpleListFilter):
    title = 'Purchase Order'
    parameter_name = 'purchase_order'

    def lookups(self, request, model_admin):
        return list(
            {
                (po_id, po_id)
                for po_id in PurchaseOrder.objects.values_list('id', flat=True)
            }
        )

    def queryset(self, request, queryset):
        product_id = request.GET.get('product_id')
        if self.value():
            return queryset.filter(
                pk__in=OrderProduct.objects.filter(
                    purchase_order_product__purchase_order__in=str(self.value()).split(
                        ','
                    ),
                    product_id__product_id=product_id,
                ).values('order_id')
            )
        return queryset


class SendableFilter(admin.SimpleListFilter):
    title = _('sendable to logistic center')
    parameter_name = 'sendable'

    def lookups(self, request, model_admin):
        return [('1', 'Yes')]

    def queryset(self, request, queryset):
        if self.value() == '1':
            # filter must be done with another query since exclude and filter
            # beehave differently -
            # https://docs.djangoproject.com/en/5.1/topics/db/queries/#spanning-multi-valued-relationships
            return queryset.filter(
                orderproduct__in=OrderProduct.objects.exclude(
                    product_id__product_id__product_type=Product.ProductTypeEnum.SENT_BY_SUPPLIER.name  # noqa: E501
                ).exclude(
                    product_id__product_id__product_kind=Product.ProductKindEnum.MONEY.name  # noqa: E501
                )
            ).distinct()


class OrderStatusFilter(MultiSelectFilter):
    custom_title = 'status'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Order.OrderStatusEnum)]


class ProductTypeFilter(MultiSelectFilter):
    custom_title = 'product type'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Product.ProductTypeEnum)]


class ProductKindFilter(MultiSelectFilter):
    custom_title = 'product kind'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Product.ProductKindEnum)]


class CampaignStatusFilter(MultiSelectFilter):
    custom_title = 'campaign status'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Campaign.CampaignStatusEnum)]


class CampaignTypeFilter(MultiSelectFilter):
    custom_title = 'campaign type'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(Campaign.CampaignTypeEnum)]


class CampaignTagFilter(MultiSelectFilter):
    custom_title = 'campaign tag'


class DeliveryTypeFilter(MultiSelectFilter):
    custom_title = 'delivery location'

    def lookups(self, request, model_admin):
        return [(nm.name, nm.value) for nm in list(DeliveryLocationEnum)]


class SupplierFilter(MultiSelectFilter):
    custom_title = 'supplier'


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin, OrderActionsMixin):
    change_list_template = 'campaign/order_change_list.html'
    list_display = (
        'order_id',
        'reference',
        'campaign',
        'organization',
        'order_date_time',
        'status',
        'po_status',
        'employee_name',
        'employee_group',
        'phone_number',
        'additional_phone_number',
        'ordered_product_names',
        'ordered_product_types',
        'ordered_product_kinds',
        'user_address',
        'dc_status',
        'dc_status_last_changed_formatted',
        'po_number',
    )
    list_filter = (
        CampaignFilter,
        OrganizationFilter,
        ('status', OrderStatusFilter),
        'order_date_time',
        ('orderproduct__product_id__product_id__product_type', ProductTypeFilter),
        ('orderproduct__product_id__product_id__product_kind', ProductKindFilter),
        ('campaign_employee_id__campaign__status', CampaignStatusFilter),
        ('campaign_employee_id__campaign__campaign_type', CampaignTypeFilter),
        ('campaign_employee_id__campaign__tags__name', CampaignTagFilter),
        (
            'campaign_employee_id__employee__employee_group__delivery_location',
            DeliveryTypeFilter,
        ),
        ('orderproduct__product_id__product_id__supplier__name', SupplierFilter),
        OrderProductFilter,
        PurchaseOrderFilter,
        ('logistics_center_status', custom_titled_filter('DC status')),
        SendableFilter,
    )
    search_fields = (
        'reference',
        'order_id',
        'orderproduct__product_id__product_id__name_en',
        'orderproduct__product_id__product_id__name_he',
        'orderproduct__product_id__product_id__product_type',
        'orderproduct__product_id__product_id__product_kind',
        'campaign_employee_id__campaign__name_en',
        'campaign_employee_id__campaign__name_he',
        'campaign_employee_id__campaign__organization__name_en',
        'campaign_employee_id__campaign__organization__name_he',
        'order_date_time',
        'status',
        'campaign_employee_id__employee__first_name_en',
        'campaign_employee_id__employee__first_name_he',
        'campaign_employee_id__employee__last_name_en',
        'campaign_employee_id__employee__last_name_he',
        'campaign_employee_id__employee__employee_group__name',
        'phone_number',
        'additional_phone_number',
        'logistics_center_status',
    )

    fieldsets = [
        (
            None,
            {
                'fields': (
                    'campaign_employee_id',
                    'order_date_time',
                    'cost_from_budget',
                    'cost_added',
                    'status',
                    'full_name',
                    'phone_number',
                    'additional_phone_number',
                    'delivery_city',
                    'delivery_street',
                    'delivery_street_number',
                    'delivery_apartment_number',
                    'delivery_additional_details',
                    'impersonated_by',
                    'logistics_center_status',
                    'country',
                    'state_code',
                    'zip_code',
                    'size',
                    'color',
                ),
            },
        ),
    ]

    form = OrderForm

    def get_readonly_fields(self, request, obj=None):
        # don't allow changing the campaign employee id for saved orders
        return ['campaign_employee_id'] if obj else []

    def get_inlines(self, request, obj):
        # only allow changing ordered products for saved orders
        return [OrderProductInline] if obj else []

    actions = ['export_as_xlsx', 'send_orders', 'complete']

    def user_address(self, obj):
        return format_with_none_replacement(
            '{delivery_street} {delivery_street_number}'
            '{delivery_apartment_number}{delivery_city}'
            '{delivery_additional_details}',
            delivery_street=obj.delivery_street,
            delivery_street_number=obj.delivery_street_number,
            delivery_apartment_number=obj.delivery_apartment_number,
            delivery_city=obj.delivery_city,
            delivery_additional_details=obj.delivery_additional_details,
        )

    def employee_group(self, obj):
        return obj.campaign_employee_id.employee.employee_group

    def dc_status(self, obj):
        return obj.logistics_center_status

    def dc_status_last_changed_formatted(self, obj):
        if obj.dc_status_last_changed:
            return obj.dc_status_last_changed.strftime('%d/%m/%Y %H:%M')
        return '-'

    def order_id(self, obj):
        # order_id is annotated by the custom model manager
        return obj.order_id

    def po_number(self, obj):
        html_content = format_html(
            ','.join(
                [
                    f'<a href="/admin/logistics/poorder/{purchase_order_id}/change/" '
                    f'target="_blank">{purchase_order_id}</a>'
                    for purchase_order_id in PurchaseOrderProduct.objects.filter(
                        id__in=obj.orderproduct_set.values_list(
                            'purchase_order_product', flat=True
                        )
                    )
                    .values_list('purchase_order__id', flat=True)
                    .distinct()  # Added distinct() in case there might be duplication.
                ]
            )
        )
        if not html_content:
            html_content = 'No PO Numbers'
        return html_content

    def po_status(self, obj):
        po_status_list = list(
            OrderProduct.objects.filter(order_id=obj).values_list(
                'purchase_order_product__purchase_order__status', flat=True
            )
        )
        status_dict = {
            'PENDING': 'WAITING',
            'SENT_TO_SUPPLIER': 'PO_SENT',
            'APPROVED': 'IN_TRANSIT',
        }
        return ','.join(
            [
                status_dict.get(item, '')
                for item in po_status_list
                if item and status_dict.get(item, '')
            ]
        )

    employee_group.short_description = 'Employee Group'
    user_address.short_description = 'User Address'
    dc_status.short_description = 'DC Status'
    po_number.short_description = 'PO Number'
    po_status.short_description = 'PO Status'
    dc_status_last_changed_formatted.short_description = 'Last Status Change'

    class Media:
        js = ('js/order_inline.js',)


@admin.register(OrganizationProduct)
class OrganizationProductAdmin(admin.ModelAdmin):
    list_display = (
        'organization',
        'product',
        'price',
    )
    list_filter = ('organization',)
    list_display_links = ('price',)


class ClientStatusListFilter(admin.SimpleListFilter):
    title = 'Client Status'
    parameter_name = 'client_status'

    def lookups(self, request, model_admin):
        # Define the available filter options
        return [
            (QuickOffer.ClientStatusEnum.READY_TO_CHECK.name, 'Ready to Check'),
            (
                QuickOffer.ClientStatusEnum.LIST_CHANGED_AFTER_APPROVE.name,
                'List Changed After Approve',
            ),
            (
                QuickOffer.ClientStatusEnum.CLIENT_ADDED_TO_LIST.name,
                'Client Added to List',
            ),
            ('--', 'No Status'),
        ]

    def queryset(self, request, queryset):
        # Filter the queryset based on the selected status
        if self.value():
            return queryset.annotate(
                _client_status=Case(
                    When(
                        Q(selected_products__isnull=False) & Q(send_my_list=True),
                        then=Value(QuickOffer.ClientStatusEnum.READY_TO_CHECK.name),
                    ),
                    When(
                        Q(selected_products__isnull=False) & Q(send_my_list=False),
                        then=Value(
                            QuickOffer.ClientStatusEnum.LIST_CHANGED_AFTER_APPROVE.name
                        ),
                    ),
                    When(
                        selected_products__isnull=False,
                        then=Value(
                            QuickOffer.ClientStatusEnum.CLIENT_ADDED_TO_LIST.name
                        ),
                    ),
                    default=Value('--'),
                    output_field=CharField(),
                )
            ).filter(_client_status=self.value())
        return queryset


@admin.register(QuickOffer)
class QuickOfferAdmin(admin.ModelAdmin, QuickOfferActionsMixin):
    list_display = [
        'quick_offer',
        'duplicate_link',
        'nicklas_status',
        'client_status',
        'last_login',
        'list_tags',
        'manager_site_link',
        'impersonate',
    ]
    actions = [
        'finish_selected_quick_offers',
        'export_selected_quick_offers_as_xlsx',
    ]
    list_filter = (
        'organization',
        'nicklas_status',
        ClientStatusListFilter,
        'tags',
        'last_login',
    )

    search_fields = (
        'name_en',
        'name_he',
        'nicklas_status',
        'last_login',
    )

    def get_queryset(self, request):
        qs = super().get_queryset(request)

        # Annotate the queryset with calculated client_status
        qs = qs.annotate(
            _client_status=Case(
                When(
                    Q(selected_products__isnull=False) & Q(send_my_list=True),
                    then=Value(QuickOffer.ClientStatusEnum.READY_TO_CHECK.name),
                ),
                When(
                    Q(selected_products__isnull=False) & Q(send_my_list=False),
                    then=Value(
                        QuickOffer.ClientStatusEnum.LIST_CHANGED_AFTER_APPROVE.name
                    ),
                ),
                When(
                    selected_products__isnull=False,
                    then=Value(QuickOffer.ClientStatusEnum.CLIENT_ADDED_TO_LIST.name),
                ),
                default=Value('--'),
                output_field=CharField(),
            )
        ).distinct()

        return qs

    def client_status(self, obj: QuickOffer):
        return obj.client_status

    def quick_offer(self, obj):
        return mark_safe(
            '<a href="%s?quick_offer_id=%s" target="_blank">%s</a>'
            % (reverse('admin:campaign_quickoffer_add'), obj.id, obj.name)
        )

    def manager_site_link(self, obj: QuickOffer):
        return format_html(
            '<a href="#" onclick="copyToClipboard(\'{0}\');'
            'return false;">Copy link</a>'
            '<input type="hidden" class="copy-link-input" value="{0}">',
            obj.manager_site_link,
        )

    def duplicate_link(self, obj):
        return mark_safe(
            '<a href="%s?quick_offer_id=%s&duplicate=1" target="_blank">duplicate</a>'
            % (reverse('admin:campaign_quickoffer_add'), obj.id)
        )

    def list_tags(self, obj):
        return ', '.join([tag.name for tag in obj.tags.all()])

    def impersonate(self, obj):
        impersonate_url = reverse('admin:campaign_impersonate_view', args=[obj.id])
        if obj.status == QuickOffer.StatusEnum.ACTIVE.name:
            impersonate_button = f"""
            <input type="submit" value="Impersonate" 
            onclick="window.open('{impersonate_url}','_blank')">
            """
        else:
            impersonate_button = ''
        return format_html(impersonate_button)

    impersonate.short_description = 'Impersonate'
    list_tags.short_description = 'Tags'

    def get_urls(self):
        app_label = self.opts.app_label
        model_name = self.opts.model_name
        return [
            path(
                'add/',
                self.admin_site.admin_view(QuickOfferCreationWizard.as_view()),
                name=f'{app_label}_{model_name}_add',
            ),
            path(
                '<int:quick_offer_id>/quick_offer_impersonate',
                QuickOfferImpersonateView.as_view(),
                name='campaign_impersonate_view',
            ),
        ] + super().get_urls()

    class Media:
        js = ('js/admin.js',)


@admin.register(QuickOfferTag)
class QuickOfferTagAdmin(admin.ModelAdmin):
    def has_module_permission(self, request):
        return False


@admin.register(PurchaseOrderProduct)
class PurchaseOrderProductAdmin(admin.ModelAdmin):
    """Admin for PurchaseOrderProduct model.
    Enables search by purchase order ID and displays purchase order
    and product fields in the list view.
    """

    search_fields = ['purchase_order__id']
    list_display = ['purchase_order', 'product_id']

    def get_search_results(self, request, queryset, search_term):
        """Filter PurchaseOrderProduct queryset by purchase order ID.

        This override performs a case-insensitive containment match on
        purchase_order.id when a search term is provided.

        Args:
            request (HttpRequest): The current request object.
            queryset (QuerySet): Initial unfiltered queryset.
            search_term (str): The term to filter by.

        Returns:
            Tuple[QuerySet, bool]: Filtered queryset and a boolean
            indicating whether the results should be marked distinct.
        """
        if search_term:
            queryset = queryset.filter(Q(purchase_order__id__icontains=search_term))
        return queryset, False

    def has_module_permission(self, request):
        return False


@admin.register(OrderProduct)
class OrderProductAdmin(admin.ModelAdmin):
    search_fields = ['purchase_order_product__purchase_order__id', 'product_id__name']
    list_display = ['purchase_order_product', 'product_id', 'quantity']

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(
            request, queryset, search_term
        )
        return queryset, use_distinct

    def has_module_permission(self, request):
        return False

    form = OrderProductForm


@admin.register(EmployeeGroupCampaignProduct)
class EmployeeGroupCampaignProductAdmin(admin.ModelAdmin):
    search_fields = ['product_id__name_he', 'product_id__sku']
    list_display = ['product_id', 'employee_group_campaign_id']

    def get_search_results(self, request, queryset, search_term):
        # Get campaign and employee group from the request
        campaign_id = request.GET.get('campaign_id')
        employee_group_id = request.GET.get('employee_group_id')

        # Apply campaign and employee group filters
        if campaign_id and employee_group_id:
            queryset = queryset.filter(
                employee_group_campaign_id__campaign_id=campaign_id,
                employee_group_campaign_id__employee_group_id=employee_group_id,
                active=True,
            )

        # Apply search term filter
        if search_term:
            queryset = queryset.filter(
                Q(product_id__name_he__icontains=search_term)
                | Q(product_id__sku__icontains=search_term)
            )

        return queryset, False

    def has_module_permission(self, request):
        return False
