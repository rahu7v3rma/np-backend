from datetime import datetime
from typing import Any

from django import forms
from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.db import models, transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.datastructures import MultiValueDict
from modeltranslation.admin import TranslationAdmin
from openpyxl import Workbook, load_workbook

from inventory.models import (
    ColorVariation,
    ProductColorVariationImage,
    ProductTextVariation,
    ProductVariation,
    TextVariation,
    Variation,
)


class XlsxImportForm(forms.Form):
    xlsx_file = forms.FileField(widget=forms.FileInput(attrs={'accept': '.xlsx'}))


class ImportableExportableAdmin(TranslationAdmin):
    """
    This is a base class for admins supporting import and export from and to
    xlsx files of their data.
    Note that only code that is relevant to *all* subclasses should live here.
    You can override functions as necessary for model-specific logic
    """

    import_form: forms.Form = XlsxImportForm
    import_related_fields: set[str] = ()
    import_multiple_value_splitter: str = '|||'
    export_fields: set[str] = ()
    import_excluded_fields: set[str] = ()
    import_excluded_fields_indexes: list = []

    def get_urls(self):
        """
        Adding a url to import data file in the existing urls.
        """
        urls = super().get_urls()
        my_urls = [
            path('import-data/', self.import_xlsx),
        ]
        return my_urls + urls

    def import_xlsx(self, request, **kwargs):
        """
        In the GET request, showing a form to submit a xlsx file
        and in POST request, creating the records in the model.
        """
        model_name = self.model._meta.model_name

        if request.method == 'POST':
            try:
                # use an atomic transaction so that no data is saved if any
                # single object failed to save
                with transaction.atomic():
                    created, updated = self.import_parse_and_save_xlsx_data(
                        kwargs, request.FILES
                    )

                    self.message_user(
                        request,
                        (
                            f'Your data file has been imported - {created} '
                            f'{model_name}(s) were created and {updated} '
                            f'{model_name}(s) were updated'
                        ),
                    )
            except RecordImportError as ex:
                for error, row_numbers in ex.errors.items():
                    self.message_user(
                        request,
                        (
                            f'Row{"s" if len(row_numbers) > 1 else ""} '
                            f'{", ".join(str(r) for r in row_numbers)} failed '
                            f'with error: {error}'
                        ),
                        level=messages.ERROR,
                    )
                return redirect(request.path)
            except Exception as ex:
                self.message_user(
                    request,
                    'An unknown error occurred: {}'.format(ex),
                    level=messages.ERROR,
                )
                return redirect(request.path)

            return redirect(
                f'admin:{self.model._meta.app_label}_{model_name}_changelist'
            )

        form = self.import_form()
        payload = {'form': form, 'model': model_name.capitalize()}
        return render(request, 'admin/import_form.html', payload)

    def import_parse_and_save_xlsx_data(
        self, extra_params: dict[str, Any], request_files: MultiValueDict
    ) -> tuple[int, int]:
        # load main xlsx file
        workbook = load_workbook(
            request_files.pop('xlsx_file')[0], data_only=True, read_only=True
        )
        worksheet = workbook.active

        # Set of static columns to identify the variation columns
        static_columns = {
            'id',
            'brand',
            'supplier',
            'reference',
            'name_en',
            'name_he',
            'product_kind',
            'voucher_type',
            'client_discount_rate',
            'supplier_discount_rate',
            'product_type',
            'description_en',
            'description_he',
            'sku',
            'link',
            'active',
            'cost_price',
            'delivery_price',
            'logistics_rate_cost_percent',
            'total_cost',
            'google_price',
            'sale_price',
            'technical_details_en',
            'technical_details_he',
            'warranty_en',
            'warranty_he',
            'exchange_value',
            'exchange_policy_en',
            'exchange_policy_he',
            'categories',
            'tags',
            'active_campaigns',
            'product_quantity',
        }

        # Read the first row as column names
        columns = [
            cell_value
            for row in worksheet.iter_rows(
                max_row=1, max_col=worksheet.max_column, values_only=True
            )
            for cell_value in row
            if cell_value
        ]

        # It will check with the help of static columns if its a variation column
        variation_columns = {}
        for idx, site_name in enumerate(columns):
            if site_name not in static_columns:
                variation = Variation.objects.filter(site_name=site_name).first()
                if not variation:
                    raise ValueError(f"Variation '{site_name}' not found")
                variation_columns[idx] = site_name
        errors = {}

        records_created = 0
        records_updated = 0

        for row_idx, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                max_col=len(columns) + 1,
                values_only=True,
            )
        ):
            # break if we hit an empty row since this should be th end of the
            # sheet
            if not row or all(v is None for v in row):
                break

            record_field_values = {}
            record_field_multi_values = {}

            try:
                shift = 0
                for col_idx, current_column in enumerate(columns):
                    if col_idx in self.import_excluded_fields_indexes:
                        shift += 1

                    if (
                        current_column not in self.import_related_fields
                        and current_column not in variation_columns.values()
                        and current_column != 'active_campaigns'
                    ):
                        parsed_value = self.import_parse_field(
                            current_column,
                            row[col_idx + shift],
                            extra_params,
                            request_files,
                        )
                        record_field_values[current_column] = parsed_value

                # the pk may or may not have been supplied
                record_pk = record_field_values.pop(self.model._meta.pk.name, None)

                if record_pk:
                    # update existing record
                    try:
                        record = self.model.objects.get(pk=record_pk)
                        for k, v in record_field_values.items():
                            setattr(record, k, v)
                        if not record.cost_price:
                            record.cost_price = 0
                        record.save()

                        records_updated += 1
                    # handle the case when the product is not found
                    except self.model.DoesNotExist:
                        record = self.model(**record_field_values)
                        if not record.cost_price:
                            record.cost_price = 0
                        record.full_clean()
                        record.save()

                        records_created += 1
                else:
                    # create record
                    record = self.model(**record_field_values)
                    if not record.cost_price:
                        record.cost_price = 0
                    record.full_clean()
                    record.save()

                    records_created += 1

                for col_idx, current_column in enumerate(columns):
                    if current_column in self.import_related_fields:
                        parsed_value = self.import_parse_related_field(
                            current_column, row[col_idx], request_files, record
                        )

                        record_field_multi_values[current_column] = parsed_value

                # add related multi-value field values
                for k, v in record_field_multi_values.items():
                    getattr(record, k).set(v)

                product_kind = row[columns.index('product_kind')]
                if product_kind.lower() == 'variation':
                    # Check if any variation column has data for this product
                    has_variation_data = any(
                        row[variation_index] for variation_index in variation_columns
                    )

                    if not has_variation_data:
                        # No variation data in any relevant column
                        # delete existing variations if exists
                        ProductVariation.objects.filter(product=record).delete()
                        # saving the product as it does not have any variation
                        if not record.cost_price:
                            record.cost_price = 0
                        record.save()
                        # Continue on next row/product
                        continue

                    for variation_index, site_name in variation_columns.items():
                        variation = Variation.objects.filter(
                            site_name=site_name
                        ).first()
                        # Extract the list of variations from the current Excel row
                        list_variations = []
                        if row[variation_index]:
                            list_variations = [
                                item.strip() for item in row[variation_index].split(',')
                            ]
                        new_values = set(list_variations)
                        # Fetch existing variations from the database
                        # for this product and variation
                        product_variation = ProductVariation.objects.filter(
                            product=record, variation=variation
                        ).first()
                        existing_values = set()
                        if product_variation:
                            if (
                                product_variation.variation.variation_kind
                                == Variation.VariationKindEnum.COLOR.name
                            ):
                                product_color_variation_images = (
                                    ProductColorVariationImage.objects.filter(
                                        product_variation=product_variation,
                                        product=record,
                                    ).all()
                                )
                                for (
                                    product_color_variation_image
                                ) in product_color_variation_images:
                                    existing_values.add(
                                        product_color_variation_image.color.name
                                    )
                            else:
                                for (
                                    text_variation
                                ) in product_variation.variation.text_variation.all():
                                    existing_values.add(text_variation.text)

                        # Find variations to add and remove
                        to_add = new_values - existing_values
                        to_remove = existing_values - new_values

                        # Add new variations
                        for value in to_add:
                            if (
                                variation.variation_kind
                                == Variation.VariationKindEnum.COLOR.name
                            ):
                                color_variation = ColorVariation.objects.filter(
                                    name=value
                                ).first()
                                if not color_variation:
                                    raise ValueError(
                                        f"Color variation '{value}' not found."
                                    )

                                try:
                                    product_variation, _ = (
                                        ProductVariation.objects.get_or_create(
                                            product=record, variation=variation
                                        )
                                    )
                                except ProductVariation.MultipleObjectsReturned:
                                    product_variation = ProductVariation.objects.filter(
                                        product=record, variation=variation
                                    ).first()
                                short_name = ProductColorVariationImage
                                try:
                                    short_name.objects.get_or_create(
                                        product_variation=product_variation,
                                        product=record,
                                        variation=variation,
                                        color=color_variation,
                                    )
                                except short_name.MultipleObjectsReturned:
                                    pass

                            elif (
                                variation.variation_kind
                                == Variation.VariationKindEnum.TEXT.name
                            ):
                                text_variation = TextVariation.objects.filter(
                                    text=value
                                ).first()
                                if not text_variation:
                                    raise ValueError(
                                        f"Text variation '{value}' not found."
                                    )

                                try:
                                    product_variation, _ = (
                                        ProductVariation.objects.get_or_create(
                                            product=record, variation=variation
                                        )
                                    )
                                except ProductVariation.MultipleObjectsReturned:
                                    product_variation = ProductVariation.objects.filter(
                                        product=record, variation=variation
                                    ).first()
                                try:
                                    ProductTextVariation.objects.get_or_create(
                                        product_variation=product_variation,
                                        product=record,
                                        variation=variation,
                                        text=text_variation,
                                    )
                                except ProductTextVariation.MultipleObjectsReturned:
                                    pass

                        # Remove stale variations
                        for value in to_remove:
                            if (
                                variation.variation_kind
                                == Variation.VariationKindEnum.COLOR.name
                            ):
                                color_variation = ColorVariation.objects.filter(
                                    name=value
                                ).first()
                                if color_variation:
                                    ProductColorVariationImage.objects.filter(
                                        product_variation__product=record,
                                        variation=variation,
                                        color=color_variation,
                                    ).delete()
                            elif (
                                variation.variation_kind
                                == Variation.VariationKindEnum.TEXT.name
                            ):
                                text_variation = TextVariation.objects.filter(
                                    text=value
                                ).first()
                                if text_variation:
                                    ProductTextVariation.objects.filter(
                                        product_variation__product=record,
                                        variation=variation,
                                        text=text_variation,
                                    ).delete()

                        # Clean up empty ProductVariation objects (if any)
                        ProductVariation.objects.filter(
                            product=record,
                            variation=variation,
                            productcolorvariationimage=None,
                            producttextvariation=None,
                        ).delete()

                if not record.cost_price:
                    record.cost_price = 0
                record.save()
            except ValidationError as ex:
                for mk in ex.message_dict.keys():
                    msg_details = ', '.join(
                        [m.removesuffix('.') for m in ex.message_dict[mk]]
                    )
                    msg = f'{mk} - {msg_details}'

                    if msg not in errors:
                        errors[msg] = []

                    # row_idx is 0-based, plus the first row is the field names
                    errors[msg].append(row_idx + 2)
            except (ValueError, Exception) as ex:
                # ValueErrors are raised by model-specific import logic (for
                # example product image matching). any other Exception should
                # also be caught here and handled the same way
                msg = str(ex)

                if msg not in errors:
                    errors[msg] = []

                # row_idx is 0-based, plus the first row is the field names
                errors[msg].append(row_idx + 2)

        if errors:
            raise RecordImportError(errors)
        else:
            return records_created, records_updated

    def import_parse_field(
        self,
        name: str,
        value: str,
        extra_params: dict[str, Any],
        extra_files: MultiValueDict,
    ):
        field_type = self.model._meta.get_field(name).get_internal_type()

        if field_type in ('CharField', 'TextField'):
            return value
        elif field_type == 'BooleanField':
            if isinstance(value, bool):
                return value
            else:
                return str(value).lower() in ('true', 'yes', '1') if value else None
        elif field_type in ('IntegerField', 'BigAutoField'):
            return int(value) if value else None
        elif field_type == 'DateField':
            if isinstance(value, datetime):
                return value
            else:
                return datetime.fromisoformat(value).date() if value else None
        elif field_type == 'FloatField':
            return float(value) if value else None
        elif field_type == 'FileField':
            return value

        raise ValueError(f'Failed to parse unknown field type: {field_type}')

    def import_parse_related_field(
        self,
        name: str,
        value: str,
        extra_files: MultiValueDict,
        main_record: models.Model,
    ):
        raise NotImplementedError(
            '`import_parse_related_field` must be implemented if '
            '`import_related_fields` are supplied to create related objects '
            'properly'
        )

    def _import_split_field_value(self, value: str):
        if not value:
            return []
        else:
            return value.split(self.import_multiple_value_splitter)

    def export_as_xlsx(self, request, queryset):
        """
        Method to write data queryset into an xlsx file.
        """
        meta = self.model._meta

        if self.export_fields:
            field_names = self.export_fields
        else:
            field_names = [field.name for field in meta.fields]

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(meta)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append([f.split('__')[0] for f in field_names])

        for obj in queryset:
            row = []

            for field_name in field_names:
                field_parts = field_name.split('__')

                value = obj

                for field_part in field_parts:
                    # for a many-relation we can join the related model's field
                    # value with the splitter
                    if type(value).__name__ == 'ManyRelatedManager':
                        value = self.import_multiple_value_splitter.join(
                            [getattr(r, field_part) for r in value.all()]
                        )
                    else:
                        value = getattr(value, field_part)

                # Handle ImageField specifically
                if hasattr(value, 'url') and 'ImageField' in str(type(value)):
                    if value:
                        value = value.url
                    else:
                        value = ''  # or some default placeholder text

                row.append(value)

            worksheet.append(row)

        workbook.save(response)

        return response

    export_as_xlsx.short_description = 'Export as XLSX'


class RecordImportError(Exception):
    errors: dict[str, list[int]]

    def __init__(self, errors):
        self.errors = errors

        super().__init__()


def custom_titled_filter(title):
    class Wrapper(admin.FieldListFilter):
        def __new__(cls, *args, **kwargs):
            instance = admin.FieldListFilter.create(*args, **kwargs)
            instance.title = title
            return instance

    return Wrapper
